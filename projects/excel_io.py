"""
Excel Import/Export for Projects

Handles bulk project import from Excel files and template generation.
Uses openpyxl. No Flask imports.
"""

import sqlite3
from datetime import datetime
from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple

from qms.core import get_logger
from qms.projects.budget import PROJECT_NUMBER_PATTERN, VALID_STAGES, parse_job_code

logger = get_logger("qms.projects.excel_io")

TEMPLATE_HEADERS = [
    "Project Name",
    "Project Number",
    "Department",
    "Manager",
    "PM Email",
    "PM Phone",
    "Stage",
    "Total Budget",
    "Weight Adjustment",
    "Start Date",
    "End Date",
    "Owner Name",
    "Owner Address",
    "Owner City",
    "Owner State",
    "Owner Zip",
    "Notes",
    "Description",
]


def generate_template() -> BytesIO:
    """Generate an Excel template for bulk project import."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Projects Import"

    for col, header in enumerate(TEMPLATE_HEADERS, 1):
        ws.cell(row=1, column=col, value=header)

    example = [
        "Example Project", "06974-230-01", "Engineering", "John Smith",
        "jsmith@example.com", "(555) 123-4567", "Course of Construction",
        50000, 1.0, "2024-01-15", "2024-12-31",
        "ABC Corporation", "123 Main St", "Springfield", "IL", "62701",
        "Sample project for reference",
        "New refrigeration system for processing facility",
    ]
    for col, value in enumerate(example, 1):
        ws.cell(row=2, column=col, value=value)

    widths = [25, 18, 15, 20, 25, 15, 22, 15, 18, 12, 12, 25, 25, 15, 8, 10, 30, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def import_projects_from_excel(
    conn: sqlite3.Connection, file_bytes: BytesIO
) -> Dict[str, Any]:
    """
    Import projects from an Excel file.

    Returns dict with success_count and errors list.
    """
    from openpyxl import load_workbook

    wb = load_workbook(file_bytes)
    ws = wb.active

    existing_codes = {
        r["number"]
        for r in conn.execute("SELECT number FROM projects").fetchall()
        if r["number"]
    }
    valid_bu_codes = {
        r["code"]
        for r in conn.execute("SELECT code FROM business_units").fetchall()
    }

    errors: List[Dict[str, Any]] = []
    valid_rows: List[Dict[str, Any]] = []
    import_codes: set = set()

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue

        project_name = str(row[0]).strip() if row[0] else ""
        project_number = str(row[1]).strip() if row[1] else ""
        department = str(row[2]).strip() if row[2] else None
        manager = str(row[3]).strip() if row[3] else None
        pm_email = str(row[4]).strip() if len(row) > 4 and row[4] else None
        pm_phone = str(row[5]).strip() if len(row) > 5 and row[5] else None
        stage = str(row[6]).strip() if len(row) > 6 and row[6] else ""
        total_budget = row[7] if len(row) > 7 else None
        weight_adj = row[8] if len(row) > 8 and row[8] is not None else 1.0
        start_date = row[9] if len(row) > 9 else None
        end_date = row[10] if len(row) > 10 else None
        owner_name = str(row[11]).strip() if len(row) > 11 and row[11] else None
        owner_addr = str(row[12]).strip() if len(row) > 12 and row[12] else None
        owner_city = str(row[13]).strip() if len(row) > 13 and row[13] else None
        owner_state = str(row[14]).strip().upper() if len(row) > 14 and row[14] else None
        owner_zip = str(row[15]).strip() if len(row) > 15 and row[15] else None
        notes = str(row[16]).strip() if len(row) > 16 and row[16] else None
        description = str(row[17]).strip() if len(row) > 17 and row[17] else None

        row_errors: List[str] = []

        if not project_name:
            row_errors.append("Project Name is required")
        if not project_number:
            row_errors.append("Project Number is required")
        elif not PROJECT_NUMBER_PATTERN.match(project_number):
            row_errors.append("Project Number must follow NNNNN or NNNNN-CCC-SS format")
        else:
            parsed = parse_job_code(project_number)
            base_number = parsed[0] if parsed else project_number
            bu_code_from_num = parsed[1] if parsed else None
            subjob_from_num = parsed[2] if parsed else None

            if bu_code_from_num and bu_code_from_num not in valid_bu_codes:
                row_errors.append(f'Business Unit code "{bu_code_from_num}" does not exist')
            if base_number in existing_codes:
                row_errors.append("Project Number already exists")
            elif base_number in import_codes:
                row_errors.append("Duplicate in import file")
            else:
                import_codes.add(base_number)

        if not stage:
            row_errors.append("Stage is required")
        elif stage not in VALID_STAGES:
            row_errors.append(f"Stage must be one of: {', '.join(VALID_STAGES)}")

        if total_budget is not None:
            try:
                total_budget = float(total_budget)
            except (ValueError, TypeError):
                row_errors.append("Total Budget must be a number")

        # Parse dates
        start_str = _parse_date_value(start_date)
        end_str = _parse_date_value(end_date)

        if row_errors:
            errors.append({
                "row": row_num,
                "project_name": project_name or "(not provided)",
                "errors": row_errors,
            })
        else:
            row_data = {
                "name": project_name,
                "code": project_number,
                "manager": manager,
                "stage": stage,
                "total_budget": total_budget or 0,
                "weight_adjustment": float(weight_adj),
                "start_date": start_str,
                "end_date": end_str,
                "notes": notes,
                "client": owner_name,
                "street": owner_addr,
                "city": owner_city,
                "state": owner_state,
                "zip_code": owner_zip,
                "description": description,
            }
            # If full code provided, create allocation for that BU
            parsed = parse_job_code(project_number)
            if parsed and parsed[1]:
                row_data["allocations"] = [{
                    "bu_code": parsed[1],
                    "subjob": parsed[2] or "00",
                    "budget": total_budget or 0,
                    "weight": float(weight_adj),
                }]
            valid_rows.append(row_data)

    from qms.projects.budget import create_project_with_budget

    success_count = 0
    for proj in valid_rows:
        try:
            create_project_with_budget(conn, **proj)
            success_count += 1
        except Exception as e:
            errors.append({
                "row": "N/A",
                "project_name": proj["name"],
                "errors": [f"Database error: {e}"],
            })

    return {"success_count": success_count, "errors": errors}


def _parse_date_value(value) -> Optional[str]:
    """Convert various date inputs to YYYY-MM-DD string."""
    if not value:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    try:
        return datetime.strptime(str(value), "%Y-%m-%d").strftime("%Y-%m-%d")
    except ValueError:
        return None
