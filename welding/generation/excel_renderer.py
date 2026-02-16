"""
Excel template renderer for welding forms.

Fills openpyxl workbook templates with data from the database,
using the form definition's template mapping.
"""

import shutil
from pathlib import Path
from typing import Any, Dict, Optional

from qms.core import get_logger
from qms.welding.forms.base import BaseFormDefinition

logger = get_logger("qms.welding.generation.excel_renderer")


def _set_cell_value(ws, cell_ref: str, value: Any):
    """Set a cell value, handling type conversion."""
    if value is None:
        return

    cell = ws[cell_ref]

    if isinstance(value, (int, float)):
        cell.value = value
    elif isinstance(value, bool):
        cell.value = "Yes" if value else "No"
    elif isinstance(value, str):
        cell.value = value
    else:
        cell.value = str(value)


def render_excel(template_path: Path, form_data: Dict[str, Any],
                 output_path: Path,
                 form_def: Optional[BaseFormDefinition] = None) -> Optional[Path]:
    """
    Fill an Excel template with form data and save to output path.

    Args:
        template_path: Path to the template Excel file.
        form_data: Dict with 'parent' and child sections from get_form_data().
        output_path: Where to save the filled workbook.
        form_def: Form definition with template_mapping. If None, uses
                  direct field-to-cell mapping from form_data keys.

    Returns:
        Path to output file, or None on failure.
    """
    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl is required for Excel generation")
        return None

    template_path = Path(template_path)
    output_path = Path(output_path)

    if not template_path.exists():
        logger.error("Template not found: %s", template_path)
        return None

    try:
        # Handle old .xls format by copying and trying to open
        if template_path.suffix.lower() == ".xls":
            logger.warning("Old .xls format detected. openpyxl requires .xlsx. "
                           "Template may need manual conversion: %s", template_path)
            # Try anyway — openpyxl can sometimes handle it
            try:
                wb = openpyxl.load_workbook(str(template_path))
            except Exception:
                logger.error("Cannot open .xls template with openpyxl. "
                             "Convert to .xlsx first: %s", template_path)
                return None
        else:
            wb = openpyxl.load_workbook(str(template_path))

        ws = wb.active

        if form_def:
            # Use form definition's template mapping
            mapping = form_def.get_template_mapping()
            parent = form_data.get("parent", {})

            for map_key, cell_ref in mapping.items():
                # Parse "table.column" format
                parts = map_key.split(".", 1)
                if len(parts) != 2:
                    continue
                table_name, column = parts

                if table_name == form_def.parent_table:
                    # Parent field
                    value = parent.get(column)
                    _set_cell_value(ws, cell_ref, value)
                else:
                    # Child field — get first row of the matching section
                    section_name = table_name.replace(f"weld_{form_def.form_type}_", "")
                    section_data = form_data.get(section_name, [])
                    if section_data and isinstance(section_data, list) and len(section_data) > 0:
                        value = section_data[0].get(column)
                        _set_cell_value(ws, cell_ref, value)
        else:
            # Direct mapping: try to place parent values in cells
            parent = form_data.get("parent", {})
            # Without a mapping, we can't know cell positions
            logger.warning("No form definition provided — skipping cell mapping")

        # Save
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output_path))
        logger.info("Excel generated: %s", output_path)
        return output_path

    except Exception as e:
        logger.error("Excel generation failed for %s: %s", template_path.name, e)
        return None


def render_excel_multi_row(template_path: Path, form_data: Dict[str, Any],
                           output_path: Path, form_def: BaseFormDefinition,
                           row_mappings: Optional[Dict[str, int]] = None) -> Optional[Path]:
    """
    Fill an Excel template with multi-row child data (e.g., multiple processes,
    multiple test results, multiple pass parameters).

    This handles the case where child tables have multiple rows that need to
    be placed in sequential Excel rows.

    Args:
        template_path: Path to template.
        form_data: Data from get_form_data().
        output_path: Output file path.
        form_def: Form definition.
        row_mappings: Optional dict mapping section_name to starting Excel row.

    Returns:
        Path to output file, or None on failure.
    """
    try:
        import openpyxl
    except ImportError:
        return None

    template_path = Path(template_path)
    if not template_path.exists():
        return None

    try:
        wb = openpyxl.load_workbook(str(template_path))
        ws = wb.active

        # First, fill parent data using the standard mapping
        mapping = form_def.get_template_mapping()
        parent = form_data.get("parent", {})

        for map_key, cell_ref in mapping.items():
            parts = map_key.split(".", 1)
            if len(parts) != 2:
                continue
            table_name, column = parts

            if table_name == form_def.parent_table:
                value = parent.get(column)
                _set_cell_value(ws, cell_ref, value)

        # Then fill multi-row sections
        if row_mappings:
            for section_name, start_row in row_mappings.items():
                section_data = form_data.get(section_name, [])
                if not section_data:
                    continue

                for i, row_data in enumerate(section_data):
                    current_row = start_row + i
                    # Place each value in columns starting from B
                    for col_idx, (key, value) in enumerate(row_data.items()):
                        if key in ("id", f"{form_def.form_type}_id"):
                            continue
                        cell = ws.cell(row=current_row, column=col_idx + 2)
                        if value is not None:
                            cell.value = value

        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output_path))
        return output_path

    except Exception as e:
        logger.error("Multi-row Excel generation failed: %s", e)
        return None
