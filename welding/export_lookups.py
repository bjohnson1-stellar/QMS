"""
Export welding lookup data to Excel for Power Automate consumption.

Writes an Excel workbook with sheets for each lookup table that
Power Automate reads to build Adaptive Card dropdowns in Teams.

The output file syncs to OneDrive automatically, so Power Automate
always has current data without requiring SharePoint or premium connectors.

Usage:
    qms welding export-lookups                      # Default output path
    qms welding export-lookups --output ~/OneDrive/QMS/welding-lookups.xlsx
    qms welding export-lookups --dry-run             # Preview row counts
"""

import sqlite3
from datetime import date
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from qms.core import get_config_value, get_db, get_logger

logger = get_logger("qms.welding.export_lookups")


# ---------------------------------------------------------------------------
# Header styling for the Excel sheets
# ---------------------------------------------------------------------------

_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)


# ---------------------------------------------------------------------------
# Sheet definitions: (sheet_name, description, column_defs, extractor)
#
# column_defs: list of (header, width, db_key)
# extractor: "db" callable(conn) -> list[dict]  OR  "static" callable() -> list[dict]
# ---------------------------------------------------------------------------

def _get_active_welders(conn: sqlite3.Connection) -> List[Dict[str, Any]]:
    """Active welders for the card dropdown."""
    rows = conn.execute(
        """SELECT welder_stamp, employee_number, display_name,
                  first_name, last_name, department, supervisor,
                  business_unit, status
           FROM weld_welder_registry
           WHERE status = 'active'
           ORDER BY last_name, first_name"""
    ).fetchall()
    return [
        {
            "welder_stamp": r["welder_stamp"] or "",
            "employee_number": r["employee_number"] or "",
            "display_name": r["display_name"] or "",
            "first_name": r["first_name"] or "",
            "last_name": r["last_name"] or "",
            "department": r["department"] or "",
            "supervisor": r["supervisor"] or "",
            "business_unit": r["business_unit"] or "",
        }
        for r in rows
    ]


def _get_active_wps(conn: sqlite3.Connection) -> List[Dict[str, Any]]:
    """Active WPS records."""
    rows = conn.execute(
        """SELECT wps_number, revision, title, status,
                  applicable_codes, is_swps
           FROM weld_wps
           WHERE status IN ('active', 'draft')
           ORDER BY wps_number"""
    ).fetchall()
    return [
        {
            "wps_number": r["wps_number"] or "",
            "revision": r["revision"] or "",
            "title": r["title"] or "",
            "status": r["status"] or "draft",
            "applicable_codes": r["applicable_codes"] or "",
            "is_swps": "Yes" if r["is_swps"] else "No",
        }
        for r in rows
    ]


def _get_field_employees(conn: sqlite3.Connection) -> List[Dict[str, Any]]:
    """Active employees with job/project assignment for jobsite filtering."""
    rows = conn.execute(
        """SELECT
               e.employee_number,
               e.first_name,
               e.last_name,
               e.preferred_name,
               COALESCE(e.preferred_name, e.first_name) || ' ' || e.last_name
                   AS display_name,
               CASE WHEN e.is_employee = 1 THEN 'Employee'
                    WHEN e.is_subcontractor = 1 THEN 'Subcontractor'
                    ELSE 'Other' END AS employee_type,
               e.position,
               d.department_number,
               d.name AS department_name,
               j.job_number,
               p.number AS project_number,
               p.name AS project_name,
               sup.first_name || ' ' || sup.last_name AS supervisor_name,
               e.status
           FROM employees e
           LEFT JOIN departments d ON e.department_id = d.id
           LEFT JOIN jobs j ON e.job_id = j.id
           LEFT JOIN projects p ON j.project_id = p.id
           LEFT JOIN employees sup ON e.supervisor_id = sup.id
           WHERE e.status = 'active' AND e.job_id IS NOT NULL
           ORDER BY p.number, e.last_name, e.first_name"""
    ).fetchall()
    return [
        {
            "employee_number": r["employee_number"] or "",
            "display_name": r["display_name"] or "",
            "first_name": r["first_name"] or "",
            "last_name": r["last_name"] or "",
            "employee_type": r["employee_type"] or "",
            "position": r["position"] or "",
            "department_number": r["department_number"] or "",
            "department_name": r["department_name"] or "",
            "job_number": r["job_number"] or "",
            "project_number": r["project_number"] or "",
            "project_name": r["project_name"] or "",
            "supervisor_name": r["supervisor_name"] or "",
        }
        for r in rows
    ]


def _get_active_projects(conn: sqlite3.Connection) -> List[Dict[str, Any]]:
    """Projects that have active field personnel assigned."""
    rows = conn.execute(
        """SELECT p.number, p.name, p.client, p.pm, p.status,
                  COUNT(e.id) AS employee_count
           FROM projects p
           JOIN jobs j ON j.project_id = p.id
           JOIN employees e ON e.job_id = j.id AND e.is_active = 1
           GROUP BY p.number, p.name, p.client, p.pm, p.status
           ORDER BY p.number"""
    ).fetchall()
    return [
        {
            "project_number": r["number"] or "",
            "project_name": r["name"] or "",
            "client": r["client"] or "",
            "pm": r["pm"] or "",
            "status": r["status"] or "",
            "employee_count": r["employee_count"],
        }
        for r in rows
    ]


def _get_wpq_status(conn: sqlite3.Connection) -> List[Dict[str, Any]]:
    """Current WPQ records with continuity status."""
    rows = conn.execute(
        """SELECT
               wpq.wpq_number,
               wr.welder_stamp,
               wr.display_name,
               wpq.process_type,
               wpq.p_number_base as p_number,
               wpq.f_number,
               wpq.groove_positions_qualified as positions_qualified,
               wpq.test_date,
               wpq.current_expiration_date,
               wpq.status,
               CAST(JULIANDAY(wpq.current_expiration_date) - JULIANDAY(DATE('now'))
                    AS INTEGER) as days_remaining,
               CASE
                   WHEN wpq.current_expiration_date < DATE('now') THEN 'LAPSED'
                   WHEN wpq.current_expiration_date < DATE('now', '+30 days') THEN 'AT_RISK'
                   ELSE 'OK'
               END as continuity_status
           FROM weld_wpq wpq
           JOIN weld_welder_registry wr ON wpq.welder_id = wr.id
           WHERE wr.status = 'active'
           ORDER BY wr.last_name, wpq.process_type"""
    ).fetchall()
    return [
        {
            "wpq_number": r["wpq_number"] or "",
            "welder_stamp": r["welder_stamp"] or "",
            "welder_name": r["display_name"] or "",
            "process_type": r["process_type"] or "",
            "p_number": r["p_number"] if r["p_number"] is not None else "",
            "f_number": r["f_number"] if r["f_number"] is not None else "",
            "positions_qualified": r["positions_qualified"] or "",
            "test_date": r["test_date"] or "",
            "expiration_date": r["current_expiration_date"] or "",
            "days_remaining": r["days_remaining"] if r["days_remaining"] is not None else "",
            "continuity_status": r["continuity_status"] or "",
            "status": r["status"] or "",
        }
        for r in rows
    ]


def _get_processes() -> List[Dict[str, Any]]:
    """Static welding process list."""
    processes = {
        "SMAW": "Shielded Metal Arc Welding",
        "GTAW": "Gas Tungsten Arc Welding",
        "GMAW": "Gas Metal Arc Welding",
        "FCAW": "Flux-Cored Arc Welding",
        "SAW": "Submerged Arc Welding",
        "GTAW/SMAW": "GTAW Root / SMAW Fill",
    }
    return [
        {"process_code": code, "process_name": name}
        for code, name in processes.items()
    ]


def _get_positions() -> List[Dict[str, Any]]:
    """Static welding position list."""
    positions = {
        "1G": "Groove", "2G": "Groove", "3G": "Groove",
        "4G": "Groove", "5G": "Groove", "6G": "Groove", "6GR": "Groove",
        "1F": "Fillet", "2F": "Fillet", "3F": "Fillet",
        "4F": "Fillet", "5F": "Fillet",
    }
    return [
        {"position": code, "type": ptype}
        for code, ptype in positions.items()
    ]


def _get_base_materials() -> List[Dict[str, Any]]:
    """Static base material reference."""
    materials = {
        "A53": (1, "Carbon Steel"),
        "A106": (1, "Carbon Steel"),
        "A333": (1, "Carbon Steel"),
        "A516": (1, "Carbon Steel"),
        "CS": (1, "Carbon Steel"),
        "A312": (8, "Stainless Steel"),
        "A358": (8, "Stainless Steel"),
        "A240": (8, "Stainless Steel"),
        "SS": (8, "Stainless Steel"),
        "SS304": (8, "Stainless Steel 304"),
        "SS316": (8, "Stainless Steel 316"),
    }
    return [
        {"specification": spec, "p_number": pnum, "material_type": mtype}
        for spec, (pnum, mtype) in materials.items()
    ]


def _get_filler_metals() -> List[Dict[str, Any]]:
    """Static filler metal reference."""
    fillers = {
        "6010": (3, "SMAW", "E6010"),
        "6011": (3, "SMAW", "E6011"),
        "7018": (4, "SMAW", "E7018"),
        "8018": (4, "SMAW", "E8018"),
        "8010": (3, "SMAW", "E8010"),
        "ER70S": (6, "GTAW", "ER70S-2/ER70S-6"),
        "ER70S2": (6, "GTAW", "ER70S-2"),
        "ER70S6": (6, "GTAW", "ER70S-6"),
        "ER80S": (6, "GTAW", "ER80S-Ni1"),
        "ER308": (6, "GTAW", "ER308L"),
        "ER309": (6, "GTAW", "ER309L"),
        "ER316": (6, "GTAW", "ER316L"),
    }
    return [
        {"filler_code": code, "f_number": fnum, "process": proc, "aws_class": aws}
        for code, (fnum, proc, aws) in fillers.items()
    ]


# ---------------------------------------------------------------------------
# Sheet configuration
# ---------------------------------------------------------------------------

SHEET_DEFS: List[Tuple[str, str, Any, List[Tuple[str, int, str]]]] = [
    (
        "Employees", "db", _get_field_employees,
        [
            ("Employee #", 14, "employee_number"),
            ("Display Name", 24, "display_name"),
            ("First Name", 14, "first_name"),
            ("Last Name", 16, "last_name"),
            ("Type", 14, "employee_type"),
            ("Position", 18, "position"),
            ("Dept #", 8, "department_number"),
            ("Department", 16, "department_name"),
            ("Job Number", 14, "job_number"),
            ("Project Number", 16, "project_number"),
            ("Project Name", 28, "project_name"),
            ("Supervisor", 22, "supervisor_name"),
        ],
    ),
    (
        "Welders", "db", _get_active_welders,
        [
            ("Welder Stamp", 14, "welder_stamp"),
            ("Employee #", 14, "employee_number"),
            ("Display Name", 24, "display_name"),
            ("First Name", 16, "first_name"),
            ("Last Name", 16, "last_name"),
            ("Department", 16, "department"),
            ("Supervisor", 20, "supervisor"),
            ("Business Unit", 16, "business_unit"),
        ],
    ),
    (
        "WPS", "db", _get_active_wps,
        [
            ("WPS Number", 14, "wps_number"),
            ("Revision", 10, "revision"),
            ("Title", 30, "title"),
            ("Status", 10, "status"),
            ("Applicable Codes", 24, "applicable_codes"),
            ("Is SWPS", 10, "is_swps"),
        ],
    ),
    (
        "Projects", "db", _get_active_projects,
        [
            ("Project Number", 16, "project_number"),
            ("Project Name", 30, "project_name"),
            ("Client", 24, "client"),
            ("PM", 20, "pm"),
            ("Status", 12, "status"),
            ("Employee Count", 16, "employee_count"),
        ],
    ),
    (
        "WPQ Status", "db", _get_wpq_status,
        [
            ("WPQ Number", 16, "wpq_number"),
            ("Welder Stamp", 14, "welder_stamp"),
            ("Welder Name", 24, "welder_name"),
            ("Process", 12, "process_type"),
            ("P-Number", 10, "p_number"),
            ("F-Number", 10, "f_number"),
            ("Positions", 16, "positions_qualified"),
            ("Test Date", 12, "test_date"),
            ("Expiration", 12, "expiration_date"),
            ("Days Left", 10, "days_remaining"),
            ("Continuity", 12, "continuity_status"),
            ("Status", 10, "status"),
        ],
    ),
    (
        "Processes", "static", _get_processes,
        [
            ("Process Code", 14, "process_code"),
            ("Process Name", 30, "process_name"),
        ],
    ),
    (
        "Positions", "static", _get_positions,
        [
            ("Position", 10, "position"),
            ("Type", 10, "type"),
        ],
    ),
    (
        "Base Materials", "static", _get_base_materials,
        [
            ("Specification", 14, "specification"),
            ("P-Number", 10, "p_number"),
            ("Material Type", 24, "material_type"),
        ],
    ),
    (
        "Filler Metals", "static", _get_filler_metals,
        [
            ("Filler Code", 12, "filler_code"),
            ("F-Number", 10, "f_number"),
            ("Process", 10, "process"),
            ("AWS Class", 20, "aws_class"),
        ],
    ),
]


# ---------------------------------------------------------------------------
# Export function
# ---------------------------------------------------------------------------

def export_lookups(
    output_path: Optional[Path] = None,
    dry_run: bool = False,
) -> Dict[str, Any]:
    """
    Export welding lookup data to an Excel workbook.

    The workbook contains one sheet per lookup table, formatted as
    Excel Tables so Power Automate can read them with the
    "List rows present in a table" action (standard connector, no premium).

    Args:
        output_path: Where to write the .xlsx file.
                     Defaults to config ``export.lookups_path`` or
                     ``data/welding-lookups.xlsx``.
        dry_run: If True, gather data but don't write the file.

    Returns:
        Dict with per-sheet row counts and the output path.
    """
    if output_path is None:
        configured = get_config_value("export", "lookups_path", default="")
        if configured:
            output_path = Path(configured)
        else:
            from qms.core.config import QMS_PATHS
            output_path = QMS_PATHS.root / "data" / "welding-lookups.xlsx"

    result: Dict[str, Any] = {
        "output_path": str(output_path),
        "sheets": {},
        "total_rows": 0,
        "exported_at": date.today().isoformat(),
    }

    # Gather all data
    sheet_data: List[Tuple[str, List[Tuple[str, int, str]], List[Dict[str, Any]]]] = []

    with get_db(readonly=True) as conn:
        for sheet_name, source_type, extractor, columns in SHEET_DEFS:
            if source_type == "db":
                rows = extractor(conn)
            else:
                rows = extractor()

            sheet_data.append((sheet_name, columns, rows))
            result["sheets"][sheet_name] = len(rows)
            result["total_rows"] += len(rows)

    if dry_run:
        result["status"] = "dry_run"
        return result

    # Build workbook
    wb = Workbook()
    default_ws = wb.active

    for sheet_name, columns, rows in sheet_data:
        ws = wb.create_sheet(title=sheet_name)

        # Write header row
        for col_idx, (header, width, _key) in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = _HEADER_ALIGN
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Write data rows
        for row_idx, row_data in enumerate(rows, start=2):
            for col_idx, (_header, _width, key) in enumerate(columns, start=1):
                ws.cell(row=row_idx, column=col_idx, value=row_data.get(key, ""))

        # Format as an Excel Table for Power Automate
        if rows:
            last_col = get_column_letter(len(columns))
            last_row = len(rows) + 1
            table_ref = f"A1:{last_col}{last_row}"
            table_name = sheet_name.replace(" ", "")
            from openpyxl.worksheet.table import Table, TableStyleInfo
            table = Table(displayName=table_name, ref=table_ref)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            ws.add_table(table)

        # Freeze header row
        ws.freeze_panes = "A2"

    # Remove the default blank sheet
    if default_ws is not None and default_ws.title == "Sheet":
        wb.remove(default_ws)

    # Write file
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))

    result["status"] = "success"
    logger.info(
        "Exported welding lookups: %d rows across %d sheets -> %s",
        result["total_rows"],
        len(sheet_data),
        output_path,
    )
    return result
