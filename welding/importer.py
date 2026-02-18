"""
Welding Excel Import

Imports welder data and WPQ qualifications from the Excel spreadsheet
(Welding Daily Log.xlsm) into the database.

Features:
    - Parses WPQ codes to extract process, P-number, position, filler metal
    - Idempotent imports via row hash comparison
    - Dry-run mode for preview
    - Progress reporting

WPQ Code Patterns (examples from actual data):
    A53-NPS6-6G-6010-7018     -> Material(A53), Size(NPS6), Position(6G), Fillers
    SS-01-P8-GTAW             -> WPS(SS-01), P-Number(8), Process(GTAW)
    CS-01-P1-SMAW             -> WPS(CS-01), P-Number(1), Process(SMAW)
"""

import hashlib
import re
import sqlite3
from datetime import date, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional

from qms.core import get_config, get_config_value, get_db, get_logger, QMS_PATHS

logger = get_logger("qms.welding.importer")


# ---------------------------------------------------------------------------
# Configuration defaults (overridable via config.yaml)
# ---------------------------------------------------------------------------

DEFAULT_EXCEL_PATH = QMS_PATHS.quality_documents / "Welding" / "Welding Daily Log.xlsm"
MAIN_SHEET = "Welder_Info"
HEADER_ROW = 3
DATA_START_ROW = 4

# Column mapping (0-indexed from Excel columns A=0)
COLUMNS: Dict[str, int] = {
    "status": 0,
    "stamp": 1,
    "employee": 2,
    "display_name": 3,
    "last_name": 4,
    "first_name": 5,
    "preferred_name": 6,
    "department": 7,
    "supervisor": 8,
    "business_unit": 9,
    "wpq1": 10,
    "wpq2": 11,
    "wpq3": 12,
    "wpq4": 13,
    "wpq5": 14,
    "running_total": 15,
    "total_tested": 16,
    "pass": 17,
    "fail": 18,
}


# ---------------------------------------------------------------------------
# Material / filler reference data
# ---------------------------------------------------------------------------

MATERIAL_P_NUMBERS: Dict[str, Optional[int]] = {
    "A53": 1, "A106": 1, "A333": 1, "A516": 1,
    "A312": 8, "A358": 8, "A240": 8,
    "SS": 8, "SS304": 8, "SS316": 8,
    "CS": 1,
    "DM": None,
}

FILLER_METAL_INFO: Dict[str, Dict[str, Any]] = {
    "6010": {"f_number": 3, "process": "SMAW", "description": "E6010"},
    "6011": {"f_number": 3, "process": "SMAW", "description": "E6011"},
    "7018": {"f_number": 4, "process": "SMAW", "description": "E7018"},
    "8018": {"f_number": 4, "process": "SMAW", "description": "E8018"},
    "8010": {"f_number": 3, "process": "SMAW", "description": "E8010"},
    "ER70S": {"f_number": 6, "process": "GTAW", "description": "ER70S-2/ER70S-6"},
    "ER70S2": {"f_number": 6, "process": "GTAW", "description": "ER70S-2"},
    "ER70S6": {"f_number": 6, "process": "GTAW", "description": "ER70S-6"},
    "ER80S": {"f_number": 6, "process": "GTAW", "description": "ER80S-Ni1"},
    "ER308": {"f_number": 6, "process": "GTAW", "description": "ER308L"},
    "ER309": {"f_number": 6, "process": "GTAW", "description": "ER309L"},
    "ER316": {"f_number": 6, "process": "GTAW", "description": "ER316L"},
}

POSITION_QUALIFIES: Dict[str, List[str]] = {
    "1G": ["1G"],
    "2G": ["1G", "2G"],
    "3G": ["1G", "3G"],
    "4G": ["1G", "4G"],
    "5G": ["1G", "2G", "5G"],
    "6G": ["1G", "2G", "3G", "4G", "5G", "6G"],
    "6GR": ["1G", "2G", "3G", "4G", "5G", "6G", "6GR"],
    "1F": ["1F"],
    "2F": ["1F", "2F"],
    "3F": ["1F", "2F", "3F"],
    "4F": ["1F", "2F", "3F", "4F"],
    "5F": ["1F", "2F", "3F", "4F", "5F"],
}


# ---------------------------------------------------------------------------
# WPQ code parser
# ---------------------------------------------------------------------------

def parse_wpq_code(code: str) -> Dict[str, Any]:
    """
    Parse a WPQ code and extract qualification details.

    Args:
        code: Raw WPQ code string (e.g. "A53-NPS6-6G-6010-7018")

    Returns:
        Dict with keys: process_type, p_number, f_number, positions,
        thickness_min, thickness_max, diameter_min, wps_number,
        raw_code, parse_notes
    """
    result: Dict[str, Any] = {
        "process_type": None,
        "p_number": None,
        "f_number": None,
        "positions": [],
        "thickness_min": None,
        "thickness_max": None,
        "diameter_min": None,
        "wps_number": None,
        "raw_code": code,
        "parse_notes": [],
    }

    if not code or code.strip() in ("", "0", "None"):
        return result

    code = code.strip().upper()
    original = code

    # Normalize: A-53 -> A53, etc.
    code = re.sub(r"^(A)[-_](\d+)", r"\1\2", code)

    # Pattern 1: WPS reference (SS-01-P8-GTAW, CS-01-P1-SMAW)
    wps_match = re.match(
        r"^(SS|CS|DM)[-_]?(\d+)[-_]?P(\d+)[-_]?(GTAW|SMAW|GMAW|FCAW)?", code, re.I
    )
    if wps_match:
        prefix, wps_num, p_num, process = wps_match.groups()
        result["wps_number"] = f"{prefix}-{wps_num}"
        result["p_number"] = int(p_num)
        if process:
            result["process_type"] = process.upper()
        if not result["process_type"]:
            if prefix == "SS":
                result["process_type"] = "GTAW"
                result["f_number"] = 6
            else:
                result["process_type"] = "SMAW"
                result["f_number"] = 4
        return result

    # Pattern 2: Material-Size-Position-Fillers (A53-NPS6-6G-6010-7018)
    mat_match = re.match(
        r"^([A-Z]+\d*)?[-_]?(NPS\d+|[\d.]+[\"']?)?[-_]?(\d+G[R]?)?[-_]?"
        r"([\dA-Z]+)?[-_]?([\dA-Z]+)?(?:[-_][\dA-Za-z]+)?",
        code,
    )
    if mat_match:
        material, size, position, filler1, filler2 = mat_match.groups()

        if material:
            material_clean = re.sub(r"[-_]", "", material)
            for mat_key, p_num in MATERIAL_P_NUMBERS.items():
                if material_clean.startswith(mat_key):
                    result["p_number"] = p_num
                    break

        if size:
            size_match = re.search(r"NPS(\d+)|(\d+\.?\d*)", size)
            if size_match:
                nps = size_match.group(1) or size_match.group(2)
                try:
                    result["diameter_min"] = float(nps)
                except ValueError:
                    pass

        if position:
            pos = position.upper()
            result["positions"] = POSITION_QUALIFIES.get(pos, [pos])

        processes: List[str] = []
        for filler in [filler1, filler2]:
            if filler:
                filler_clean = filler.upper()
                for fm_key, fm_info in FILLER_METAL_INFO.items():
                    if filler_clean.startswith(fm_key) or filler_clean == fm_key:
                        processes.append(fm_info["process"])
                        if result["f_number"] is None or fm_info["f_number"] < result["f_number"]:
                            result["f_number"] = fm_info["f_number"]
                        break

        if processes:
            unique_processes = list(set(processes))
            if len(unique_processes) == 1:
                result["process_type"] = unique_processes[0]
            elif "GTAW" in unique_processes and "SMAW" in unique_processes:
                result["process_type"] = "GTAW/SMAW"
            else:
                result["process_type"] = "/".join(unique_processes)

    # Pattern 3: Short code (1-6G-7)
    short_match = re.match(r"^(\d+)[-_](\d+G[R]?)[-_](\d+[A]?)$", code)
    if short_match:
        p_num_str, position, thickness_code = short_match.groups()
        result["p_number"] = int(p_num_str)
        result["positions"] = POSITION_QUALIFIES.get(position.upper(), [position.upper()])
        result["process_type"] = "SMAW"
        result["f_number"] = 4

        thickness_map = {
            "7": (0.75, 999),
            "8": (0.75, 999),
            "8A": (0.75, 999),
        }
        if thickness_code in thickness_map:
            result["thickness_min"], result["thickness_max"] = thickness_map[thickness_code]

    # Check for process keywords if still unknown
    if not result["process_type"]:
        code_upper = original.upper()
        if "GTAW" in code_upper:
            result["process_type"] = "GTAW"
            result["f_number"] = result["f_number"] or 6
        elif "SMAW" in code_upper:
            result["process_type"] = "SMAW"
            result["f_number"] = result["f_number"] or 4
        elif "GMAW" in code_upper:
            result["process_type"] = "GMAW"
            result["f_number"] = result["f_number"] or 6
        elif "FCAW" in code_upper:
            result["process_type"] = "FCAW"
            result["f_number"] = result["f_number"] or 6

    # Incomplete code with P-number only
    if not result["process_type"] and result["p_number"]:
        result["process_type"] = "SMAW"
        result["f_number"] = 4
        result["parse_notes"].append(f"Incomplete code, assumed SMAW: {original}")

    # SWPS references
    swps_match = re.search(r"SWPS|AWS[-\s]*(D1|B2)", original)
    if swps_match and not result["process_type"]:
        result["process_type"] = "SMAW"
        result["f_number"] = 4
        result["wps_number"] = original
        return result

    # Fallback
    if not result["process_type"]:
        result["process_type"] = "UNKNOWN"
        result["parse_notes"].append(f"Could not determine process from code: {original}")

    return result


def compute_row_hash(row: tuple) -> str:
    """Compute MD5 hash of row data for idempotent imports."""
    key_cols = [
        str(row[i]) if i < len(row) and row[i] is not None else ""
        for i in range(COLUMNS["status"], COLUMNS["fail"] + 1)
    ]
    data = "|".join(key_cols)
    return hashlib.md5(data.encode()).hexdigest()


# ---------------------------------------------------------------------------
# Database operations
# ---------------------------------------------------------------------------

def find_welder_by_stamp(conn: sqlite3.Connection, stamp: str) -> Optional[Dict[str, Any]]:
    """Find welder by stamp number."""
    row = conn.execute(
        "SELECT * FROM weld_welder_registry WHERE welder_stamp = ?", (stamp,)
    ).fetchone()
    return dict(row) if row else None


def find_welder_by_employee(conn: sqlite3.Connection, emp_num: str) -> Optional[Dict[str, Any]]:
    """Find welder by employee number."""
    row = conn.execute(
        "SELECT * FROM weld_welder_registry WHERE employee_number = ?", (emp_num,)
    ).fetchone()
    return dict(row) if row else None


def _resolve_business_unit_id(conn: sqlite3.Connection, bu_text: Optional[str]) -> Optional[int]:
    """Look up business_unit_id from the business_units table by code or name."""
    if not bu_text:
        return None
    bu_text = str(bu_text).strip()
    # Try matching by 3-digit code first
    row = conn.execute(
        "SELECT id FROM business_units WHERE code = ?", (bu_text,)
    ).fetchone()
    if row:
        return row["id"]
    # Try matching by name
    row = conn.execute(
        "SELECT id FROM business_units WHERE name = ? OR full_name = ?", (bu_text, bu_text)
    ).fetchone()
    return row["id"] if row else None


def upsert_welder(conn: sqlite3.Connection, welder_data: Dict[str, Any], row_hash: str) -> int:
    """
    Insert or update welder record.

    Args:
        conn: Database connection
        welder_data: Dict of welder fields
        row_hash: MD5 hash of source row for change detection

    Returns:
        Welder registry ID
    """
    existing = None
    if welder_data.get("welder_stamp"):
        existing = find_welder_by_stamp(conn, welder_data["welder_stamp"])
    if not existing and welder_data.get("employee_number"):
        existing = find_welder_by_employee(conn, welder_data["employee_number"])

    # Resolve business_unit_id FK from text value
    bu_id = _resolve_business_unit_id(conn, welder_data.get("business_unit"))

    if existing:
        if existing.get("excel_row_hash") != row_hash:
            conn.execute(
                """UPDATE weld_welder_registry SET
                       last_name = ?, first_name = ?, preferred_name = ?,
                       display_name = ?, department = ?, supervisor = ?,
                       business_unit = ?, business_unit_id = ?,
                       status = ?, running_total_welds = ?,
                       total_welds_tested = ?, welds_passed = ?, welds_failed = ?,
                       excel_row_hash = ?, updated_at = CURRENT_TIMESTAMP
                   WHERE id = ?""",
                (
                    welder_data.get("last_name"),
                    welder_data.get("first_name"),
                    welder_data.get("preferred_name"),
                    welder_data.get("display_name"),
                    welder_data.get("department"),
                    welder_data.get("supervisor"),
                    welder_data.get("business_unit"),
                    bu_id,
                    welder_data.get("status", "active"),
                    welder_data.get("running_total_welds", 0),
                    welder_data.get("total_welds_tested", 0),
                    welder_data.get("welds_passed", 0),
                    welder_data.get("welds_failed", 0),
                    row_hash,
                    existing["id"],
                ),
            )
            conn.commit()
        return existing["id"]

    emp_num = (
        welder_data.get("employee_number")
        or welder_data.get("welder_stamp")
        or f"UNK-{row_hash[:8]}"
    )

    cursor = conn.execute(
        """INSERT INTO weld_welder_registry (
               employee_number, welder_stamp, last_name, first_name,
               preferred_name, display_name, department, supervisor,
               business_unit, business_unit_id, status,
               running_total_welds, total_welds_tested,
               welds_passed, welds_failed, excel_row_hash
           ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        (
            emp_num,
            welder_data.get("welder_stamp"),
            welder_data.get("last_name"),
            welder_data.get("first_name"),
            welder_data.get("preferred_name"),
            welder_data.get("display_name"),
            welder_data.get("department"),
            welder_data.get("supervisor"),
            welder_data.get("business_unit"),
            bu_id,
            welder_data.get("status", "active"),
            welder_data.get("running_total_welds", 0),
            welder_data.get("total_welds_tested", 0),
            welder_data.get("welds_passed", 0),
            welder_data.get("welds_failed", 0),
            row_hash,
        ),
    )
    conn.commit()
    return cursor.lastrowid


def upsert_wpq(
    conn: sqlite3.Connection, welder_id: int, wpq_code: str, parsed: Dict[str, Any]
) -> Optional[int]:
    """
    Insert or update WPQ record from parsed code.

    Returns:
        WPQ ID or None if no valid data
    """
    if not parsed.get("process_type") or parsed["process_type"] == "UNKNOWN":
        return None

    row = conn.execute(
        "SELECT welder_stamp FROM weld_welder_registry WHERE id = ?", (welder_id,)
    ).fetchone()
    stamp = row["welder_stamp"] if row else "UNK"

    wps = parsed.get("wps_number")
    if wps:
        wpq_number = f"{stamp}-{wps}"
    else:
        wpq_number = f"{stamp}-{wpq_code}"

    existing = conn.execute(
        "SELECT id FROM weld_wpq WHERE wpq_number = ?", (wpq_number,)
    ).fetchone()

    initial_expiration = date.today() + timedelta(days=180)

    positions_str = ", ".join(parsed.get("positions", [])) if parsed.get("positions") else None

    if existing:
        conn.execute(
            """UPDATE weld_wpq SET
                   wps_number = COALESCE(?, wps_number),
                   process_type = ?,
                   p_number_base = COALESCE(?, p_number_base),
                   f_number = COALESCE(?, f_number),
                   groove_positions_qualified = COALESCE(?, groove_positions_qualified),
                   thickness_qualified_min = COALESCE(?, thickness_qualified_min),
                   thickness_qualified_max = COALESCE(?, thickness_qualified_max),
                   diameter_qualified_min = COALESCE(?, diameter_qualified_min),
                   updated_at = CURRENT_TIMESTAMP
               WHERE id = ?""",
            (
                parsed.get("wps_number"),
                parsed["process_type"],
                parsed.get("p_number"),
                parsed.get("f_number"),
                positions_str,
                parsed.get("thickness_min"),
                parsed.get("thickness_max"),
                parsed.get("diameter_min"),
                existing["id"],
            ),
        )
        conn.commit()
        return existing["id"]

    cursor = conn.execute(
        """INSERT INTO weld_wpq (
               wpq_number, welder_id, welder_stamp, wps_number,
               process_type, p_number_base, f_number,
               groove_positions_qualified, thickness_qualified_min,
               thickness_qualified_max, diameter_qualified_min,
               initial_expiration_date, current_expiration_date,
               status, notes
           ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'active', ?)""",
        (
            wpq_number,
            welder_id,
            stamp,
            parsed.get("wps_number"),
            parsed["process_type"],
            parsed.get("p_number"),
            parsed.get("f_number"),
            positions_str,
            parsed.get("thickness_min"),
            parsed.get("thickness_max"),
            parsed.get("diameter_min"),
            initial_expiration,
            initial_expiration,
            f"Imported from Excel code: {wpq_code}",
        ),
    )
    conn.commit()
    return cursor.lastrowid


# ---------------------------------------------------------------------------
# Import entry points
# ---------------------------------------------------------------------------

def get_excel_path() -> Path:
    """Get Excel path from config or use default."""
    return Path(
        get_config_value("welding", "excel_path", default=str(DEFAULT_EXCEL_PATH))
    )


def validate_excel(excel_path: Optional[Path] = None) -> Dict[str, Any]:
    """
    Validate Excel file structure.

    Args:
        excel_path: Path to Excel file (uses config default if None)

    Returns:
        Dict with valid (bool), headers (list), data_count (int), errors (list)
    """
    try:
        import openpyxl
    except ImportError:
        return {"valid": False, "errors": ["openpyxl not installed. Run: pip install openpyxl"]}

    path = excel_path or get_excel_path()
    result: Dict[str, Any] = {"valid": False, "headers": [], "data_count": 0, "errors": []}

    if not path.exists():
        result["errors"].append(f"Excel file not found: {path}")
        return result

    try:
        wb = openpyxl.load_workbook(str(path), data_only=True)
    except Exception as exc:
        result["errors"].append(f"Cannot open Excel file: {exc}")
        return result

    if MAIN_SHEET not in wb.sheetnames:
        result["errors"].append(f"Sheet '{MAIN_SHEET}' not found. Available: {wb.sheetnames}")
        return result

    ws = wb[MAIN_SHEET]
    headers = [ws.cell(row=HEADER_ROW, column=i + 1).value for i in range(7)]
    result["headers"] = headers

    expected = ["Status", "Stamp Number", "Employee #", "Display Name",
                "Last Name", "First Name", "Preferred Name"]

    for i, (exp, actual) in enumerate(zip(expected, headers)):
        if exp.lower() != (actual or "").lower():
            result["errors"].append(
                f"Column {i + 1} header mismatch: expected '{exp}', got '{actual}'"
            )

    data_count = 0
    for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        if row[0]:
            data_count += 1
    result["data_count"] = data_count

    if not result["errors"]:
        result["valid"] = True

    logger.info("Validation %s: %d data rows", "PASSED" if result["valid"] else "FAILED", data_count)
    return result


def import_from_excel(
    dry_run: bool = False,
    single_welder: Optional[str] = None,
    excel_path: Optional[Path] = None,
) -> Dict[str, Any]:
    """
    Import welders and WPQs from Excel.

    Args:
        dry_run: If True, report changes without modifying database
        single_welder: Import only this welder (by stamp or employee number)
        excel_path: Path to Excel file (uses config default if None)

    Returns:
        Dict with import statistics
    """
    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl not installed. Run: pip install openpyxl")
        return {"errors": ["openpyxl not installed"]}

    path = excel_path or get_excel_path()
    stats: Dict[str, Any] = {
        "welders_processed": 0,
        "welders_created": 0,
        "welders_updated": 0,
        "welders_skipped": 0,
        "wpqs_created": 0,
        "wpqs_updated": 0,
        "wpq_parse_errors": 0,
        "errors": [],
    }

    logger.info("Loading: %s", path)
    wb = openpyxl.load_workbook(str(path), data_only=True)
    ws = wb[MAIN_SHEET]

    conn = None
    if not dry_run:
        # Use get_db context manager, but we need the connection for the full loop
        db_ctx = get_db()
        conn = db_ctx.__enter__()

    try:
        for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
            if not row[COLUMNS["status"]]:
                continue

            status = str(row[COLUMNS["status"]]).strip().lower()
            stamp = str(row[COLUMNS["stamp"]]) if row[COLUMNS["stamp"]] else None
            employee = str(row[COLUMNS["employee"]]) if row[COLUMNS["employee"]] else None

            if not stamp and not employee:
                stats["welders_skipped"] += 1
                continue

            if stamp and stamp not in ("None", "0", ""):
                stamp = stamp.strip()
            else:
                stamp = None

            if single_welder:
                if stamp != single_welder and employee != single_welder:
                    continue

            stats["welders_processed"] += 1

            welder_data: Dict[str, Any] = {
                "status": "active" if status == "active" else "inactive",
                "welder_stamp": stamp,
                "employee_number": (
                    employee
                    if employee not in ("None", "0", "Rig Welder", "ULG", "NCW")
                    else stamp
                ),
                "last_name": str(row[COLUMNS["last_name"]]) if row[COLUMNS["last_name"]] else "",
                "first_name": str(row[COLUMNS["first_name"]]) if row[COLUMNS["first_name"]] else "",
                "preferred_name": (
                    str(row[COLUMNS["preferred_name"]]) if row[COLUMNS["preferred_name"]] else None
                ),
                "display_name": (
                    str(row[COLUMNS["display_name"]]) if row[COLUMNS["display_name"]] else None
                ),
                "department": str(row[COLUMNS["department"]]) if row[COLUMNS["department"]] else None,
                "supervisor": str(row[COLUMNS["supervisor"]]) if row[COLUMNS["supervisor"]] else None,
                "business_unit": (
                    str(row[COLUMNS["business_unit"]]) if row[COLUMNS["business_unit"]] else None
                ),
                "running_total_welds": (
                    int(row[COLUMNS["running_total"]] or 0)
                    if COLUMNS["running_total"] < len(row) else 0
                ),
                "total_welds_tested": (
                    int(row[COLUMNS["total_tested"]] or 0)
                    if COLUMNS["total_tested"] < len(row) else 0
                ),
                "welds_passed": (
                    int(row[COLUMNS["pass"]] or 0) if COLUMNS["pass"] < len(row) else 0
                ),
                "welds_failed": (
                    int(row[COLUMNS["fail"]] or 0) if COLUMNS["fail"] < len(row) else 0
                ),
            }

            # Clean "None" strings
            for key in welder_data:
                if welder_data[key] in ("None", "none", ""):
                    welder_data[key] = None

            row_hash = compute_row_hash(row)

            if stats["welders_processed"] % 50 == 0:
                logger.info("Processed %d welders...", stats["welders_processed"])

            if dry_run:
                logger.info(
                    "[DRY] %s: %s (%s)",
                    welder_data["welder_stamp"] or welder_data["employee_number"],
                    welder_data["display_name"],
                    welder_data["status"],
                )
            else:
                existing = None
                if welder_data["welder_stamp"]:
                    existing = find_welder_by_stamp(conn, welder_data["welder_stamp"])

                welder_id = upsert_welder(conn, welder_data, row_hash)

                if existing:
                    if existing.get("excel_row_hash") != row_hash:
                        stats["welders_updated"] += 1
                else:
                    stats["welders_created"] += 1

            # Process WPQ codes (columns K-O)
            wpq_codes: List[str] = []
            for i in range(COLUMNS["wpq1"], COLUMNS["wpq5"] + 1):
                if i < len(row) and row[i]:
                    code = str(row[i]).strip()
                    if code and code not in ("", "0", "None") and not code.isdigit():
                        wpq_codes.append(code)

            for wpq_code in wpq_codes:
                parsed = parse_wpq_code(wpq_code)

                if parsed.get("parse_notes"):
                    stats["wpq_parse_errors"] += 1
                    stats["errors"].append(
                        f"{welder_data.get('welder_stamp')}: {wpq_code} - {parsed['parse_notes']}"
                    )

                if dry_run:
                    logger.info(
                        "    WPQ: %s -> %s P%s F%s %s",
                        wpq_code,
                        parsed["process_type"],
                        parsed.get("p_number"),
                        parsed.get("f_number"),
                        parsed.get("positions", []),
                    )
                else:
                    wpq_id = upsert_wpq(conn, welder_id, wpq_code, parsed)
                    if wpq_id:
                        stats["wpqs_created"] += 1
    finally:
        if conn and not dry_run:
            db_ctx.__exit__(None, None, None)

    logger.info(
        "Import complete: %d processed, %d created, %d updated",
        stats["welders_processed"],
        stats["welders_created"],
        stats["welders_updated"],
    )
    return stats
