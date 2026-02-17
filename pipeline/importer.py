"""
Pipeline Drawing Importer

Handles single drawing import and batch/bulk import from directories.
Reads parsed SIS Excel workbooks or raw SIS sheets, cross-references
against the welder registry, and creates production weld / continuity
records.

Combines the functionality of the legacy sis_import.py (single import)
and sis_bulk_import.py (batch processing) into one module.
"""

import re
import sqlite3
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from qms.core import get_config, get_db, get_logger

from .common import (
    extract_date_from_filename,
    extract_project_number,
    normalize_job_numbers,
    parse_date,
)

logger = get_logger("qms.pipeline.importer")


# ---------------------------------------------------------------------------
# Excel reading (parsed workbook format — Jobsites + Field Personnel sheets)
# ---------------------------------------------------------------------------

def _load_workbook(filepath: Path):
    """Load Excel workbook via openpyxl."""
    import openpyxl
    return openpyxl.load_workbook(str(filepath), data_only=True)


def read_jobsites_sheet(wb) -> List[Dict[str, str]]:
    """
    Read 'Jobsites' sheet from a parsed workbook.

    Expects columns: Job #, Project Name, PM, Address, City, State, Zip.

    Args:
        wb: openpyxl Workbook object.

    Returns:
        List of jobsite dicts.
    """
    ws = wb["Jobsites"] if "Jobsites" in wb.sheetnames else None
    if not ws:
        logger.warning("No 'Jobsites' sheet found in workbook")
        return []

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []

    results = []
    for row in rows[1:]:
        if not row[0]:
            continue
        results.append({
            'job_number': str(row[0]).strip(),
            'project_name': str(row[1]).strip() if row[1] else '',
            'pm': str(row[2]).strip() if row[2] else '',
            'street': str(row[3]).strip() if row[3] else '',
            'city': str(row[4]).strip() if row[4] else '',
            'state': str(row[5]).strip() if row[5] else '',
            'zip': str(row[6]).strip() if row[6] else '',
        })
    return results


def read_personnel_sheet(wb) -> List[Dict[str, str]]:
    """
    Read 'Field Personnel' sheet from a parsed workbook.

    Expects columns: EMPL #, Last Name, First Name, Job #, Phone Number, Designation.

    Args:
        wb: openpyxl Workbook object.

    Returns:
        List of personnel dicts.
    """
    ws = wb["Field Personnel"] if "Field Personnel" in wb.sheetnames else None
    if not ws:
        logger.warning("No 'Field Personnel' sheet found in workbook")
        return []

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []

    results = []
    for row in rows[1:]:
        empl = str(row[0]).strip() if row[0] else ''
        if not empl:
            continue
        results.append({
            'employee_number': empl,
            'last_name': str(row[1]).strip() if row[1] else '',
            'first_name': str(row[2]).strip() if row[2] else '',
            'job_number': str(row[3]).strip() if row[3] else '',
            'phone': str(row[4]).strip() if row[4] else '',
            'designation': str(row[5]).strip() if row[5] else '',
        })
    return results


# ---------------------------------------------------------------------------
# Project / jobsite upsert helpers
# ---------------------------------------------------------------------------

def upsert_project(
    conn: sqlite3.Connection,
    project_number: str,
    project_name: str,
) -> Optional[int]:
    """
    Look up an existing project by 5-digit number.

    Projects must be created via the Projects UI or Procore import first.
    Returns None if the project doesn't exist.

    Args:
        conn: Database connection.
        project_number: 5-digit project prefix.
        project_name: Human-readable project name (used for logging only).

    Returns:
        Project database ID, or None if not found.
    """
    cursor = conn.execute(
        "SELECT id FROM projects WHERE number = ?", (project_number,)
    )
    row = cursor.fetchone()
    if row:
        return row['id']

    logger.warning(
        "Project %s (%s) not found — create it in the Projects page first",
        project_number, project_name,
    )
    return None


def upsert_jobsite(
    conn: sqlite3.Connection,
    jobsite: Dict[str, str],
    project_id: int,
    import_date: date,
) -> None:
    """
    Legacy jobsite upsert — the jobsites table has been dropped.

    Now a no-op. Job/PM data is handled by processor.py's upsert_job().
    Kept for API compatibility with existing callers.
    """
    pass


# ---------------------------------------------------------------------------
# Welder matching
# ---------------------------------------------------------------------------

def find_welder(
    conn: sqlite3.Connection, employee_number: str
) -> Optional[Dict[str, Any]]:
    """
    Find welder by employee number in the welder registry.

    Args:
        conn: Database connection.
        employee_number: Employee number to look up.

    Returns:
        Welder dict or None if not found.
    """
    cursor = conn.execute(
        "SELECT id, employee_number, welder_stamp, display_name, status "
        "FROM weld_welder_registry WHERE employee_number = ?",
        (employee_number,)
    )
    row = cursor.fetchone()
    return dict(row) if row else None


def get_active_wpq_processes(
    conn: sqlite3.Connection, welder_id: int
) -> List[str]:
    """
    Get distinct active WPQ process types for a welder.

    Args:
        conn: Database connection.
        welder_id: Welder registry ID.

    Returns:
        List of process type strings (e.g. ['GTAW', 'SMAW']).
    """
    cursor = conn.execute(
        "SELECT DISTINCT process_type FROM weld_wpq "
        "WHERE welder_id = ? AND status = 'active'",
        (welder_id,)
    )
    return [row['process_type'] for row in cursor.fetchall()]


def add_production_weld(
    conn: sqlite3.Connection,
    welder_id: int,
    project_number: str,
    process: str,
    weld_date: date,
    week_ending: date,
) -> None:
    """
    Insert production weld record.

    The tr_production_weld_continuity trigger automatically extends
    WPQ current_expiration_date and creates weld_continuity_log entries.

    Args:
        conn: Database connection.
        welder_id: Welder registry ID.
        project_number: Job/project number.
        process: Welding process type.
        weld_date: Date of the weld.
        week_ending: Week ending date for batch imports.
    """
    conn.execute("""
        INSERT INTO weld_production_welds (
            welder_id, project_number, process_type, weld_date,
            week_ending, counts_for_continuity, created_by
        ) VALUES (?, ?, ?, ?, ?, 1, 'sis_import')
        ON CONFLICT(welder_id, project_number, week_ending) DO UPDATE SET
            process_type = excluded.process_type,
            weld_date = excluded.weld_date
    """, (welder_id, project_number, process, weld_date, week_ending))


# ---------------------------------------------------------------------------
# Single drawing import (from parsed Jobsites/Field Personnel workbook)
# ---------------------------------------------------------------------------

def import_single(
    filepath: Path,
    week_override: Optional[date] = None,
    preview: bool = False,
) -> Dict[str, Any]:
    """
    Import a single SIS Excel workbook (parsed Jobsites + Field Personnel format).

    1. Read Jobsites + Field Personnel from Excel
    2. Upsert jobsites (with project linkage)
    3. Cross-reference personnel against welder registry
    4. Create production welds for matched welders

    Args:
        filepath: Path to the Excel file.
        week_override: Override the week ending date (default: today).
        preview: If True, read and report without making changes.

    Returns:
        Statistics dict with counts for all import operations.
    """
    stats: Dict[str, Any] = {
        'jobsites_created': 0,
        'jobsites_updated': 0,
        'projects_created': 0,
        'welders_matched': 0,
        'welders_not_found': [],
        'non_welders': 0,
        'production_welds_created': 0,
        'wpq_processes_extended': 0,
        'personnel_no_job': 0,
        'errors': [],
    }

    weld_date = week_override or date.today()
    week_ending = weld_date

    wb = _load_workbook(filepath)
    jobsite_rows = read_jobsites_sheet(wb)
    personnel_rows = read_personnel_sheet(wb)

    logger.info("Read %d jobsites, %d personnel from %s",
                len(jobsite_rows), len(personnel_rows), filepath.name)

    if preview:
        logger.info("PREVIEW MODE - No changes will be made")
        for js in jobsite_rows:
            proj_num = extract_project_number(js['job_number'])
            logger.info("[PREVIEW] Jobsite %s -> project %s | %s | PM: %s",
                        js['job_number'], proj_num, js['project_name'], js['pm'])
        for person in personnel_rows:
            logger.info("[PREVIEW] %s %s, %s @ %s",
                        person['employee_number'], person['last_name'],
                        person['first_name'], person['job_number'])
        stats['jobsites_processed'] = len(jobsite_rows)
        stats['personnel_processed'] = len(personnel_rows)
        return stats

    with get_db() as conn:
        # Upsert jobsites
        seen_projects: set = set()
        for js in jobsite_rows:
            proj_num = extract_project_number(js['job_number'])

            project_id = upsert_project(conn, proj_num, js['project_name'])
            if project_id is None:
                stats.setdefault('projects_skipped', set()).add(proj_num)
                continue

            # Legacy: jobsites table was dropped; track stats for compat
            stats['jobsites_updated'] += 1
            upsert_jobsite(conn, js, project_id, weld_date)
        conn.commit()

        # Cross-reference personnel, create production welds
        for person in personnel_rows:
            empl = person['employee_number']
            job = person['job_number']

            if not job:
                stats['personnel_no_job'] += 1
                continue

            welder = find_welder(conn, empl)
            if not welder:
                stats['non_welders'] += 1
                continue

            if welder['status'] != 'active':
                stats['non_welders'] += 1
                continue

            stats['welders_matched'] += 1
            processes = get_active_wpq_processes(conn, welder['id'])

            if not processes:
                continue

            for proc in processes:
                try:
                    add_production_weld(conn, welder['id'], job, proc,
                                        weld_date, week_ending)
                    stats['production_welds_created'] += 1
                    stats['wpq_processes_extended'] += 1
                    logger.info("[OK] %s: %s | %s | week %s",
                                welder.get('welder_stamp', empl),
                                job, proc, week_ending)
                except Exception as e:
                    stats['errors'].append(f"{empl}: {str(e)}")
                    logger.error("Error adding production weld: %s", e)

        conn.commit()

    return stats


# ---------------------------------------------------------------------------
# Batch / bulk import
# ---------------------------------------------------------------------------

def find_field_location_files(
    directory: Path,
) -> List[Tuple[Path, date]]:
    """
    Find all field location Excel files in a directory.

    Scans for .xlsx and .xls files, extracts dates from filenames,
    and returns them sorted chronologically (oldest first).

    Args:
        directory: Directory to scan.

    Returns:
        List of (file_path, week_ending_date) tuples, sorted by date.
    """
    files_with_dates: List[Tuple[Path, date]] = []

    for pattern in ['*.xlsx', '*.xls']:
        for file_path in directory.glob(pattern):
            if file_path.name.startswith('~'):
                continue

            week_date = extract_date_from_filename(file_path)

            if week_date:
                files_with_dates.append((file_path, week_date))
            else:
                logger.warning("Could not extract date from %s, skipping", file_path.name)

    files_with_dates.sort(key=lambda x: x[1])
    return files_with_dates


def import_batch(
    files: List[Tuple[Path, date]],
    preview: bool = False,
) -> Dict[str, Any]:
    """
    Import multiple field location files chronologically.

    Processes files oldest-to-newest so that hire dates, job changes,
    and continuity records are built up in the correct order.

    Args:
        files: List of (file_path, week_ending_date) tuples.
        preview: If True, run in preview mode (no database changes).

    Returns:
        Aggregate statistics dict across all files.
    """
    # Lazy import to avoid circular dependency — processor uses importer indirectly
    from .processor import process_and_import

    total_stats: Dict[str, Any] = {
        'files_processed': 0,
        'files_failed': 0,
        'total_jobsites': 0,
        'total_personnel': 0,
        'total_employees_created': 0,
        'total_employees_updated': 0,
        'total_job_changes': 0,
        'errors': [],
    }

    for i, (file_path, week_date) in enumerate(files, 1):
        logger.info("Processing file %d/%d: %s (week ending: %s)",
                     i, len(files), file_path.name, week_date)

        try:
            stats = process_and_import(
                file_path,
                week_override=week_date,
                output_path=None,
                preview=preview,
            )

            total_stats['files_processed'] += 1
            total_stats['total_jobsites'] += stats.get('jobsites_processed', 0)
            total_stats['total_personnel'] += stats.get('personnel_processed', 0)
            total_stats['total_employees_created'] += stats.get('employees_created', 0)
            total_stats['total_employees_updated'] += stats.get('employees_updated', 0)

            if stats.get('employees_updated', 0) > 0 and stats.get('employees_job_assigned', 0) > 0:
                job_change_estimate = min(
                    stats['employees_updated'], stats['employees_job_assigned']
                )
                total_stats['total_job_changes'] += job_change_estimate

            logger.info("File %d/%d completed successfully", i, len(files))

        except Exception as e:
            logger.error("ERROR processing file %d: %s", i, e)
            total_stats['files_failed'] += 1
            total_stats['errors'].append({
                'file': file_path.name,
                'week': str(week_date),
                'error': str(e),
            })

    return total_stats


def import_from_directory(
    directory: Path,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    preview: bool = False,
) -> Dict[str, Any]:
    """
    Scan a directory and batch-import all field location files.

    Convenience wrapper that finds files, optionally filters by date range,
    and runs import_batch.

    Args:
        directory: Directory containing field location Excel files.
        start_date: Only include files on or after this date.
        end_date: Only include files on or before this date.
        preview: If True, run in preview mode (no database changes).

    Returns:
        Aggregate statistics dict.
    """
    if not directory.exists():
        raise FileNotFoundError(f"Directory not found: {directory}")

    logger.info("Scanning directory: %s", directory)
    files_with_dates = find_field_location_files(directory)

    if not files_with_dates:
        logger.warning("No field location files found with recognizable dates")
        return {'files_processed': 0, 'files_failed': 0, 'errors': []}

    # Apply date filters
    if start_date or end_date:
        filtered = []
        for file_path, week_date in files_with_dates:
            if start_date and week_date < start_date:
                continue
            if end_date and week_date > end_date:
                continue
            filtered.append((file_path, week_date))
        files_with_dates = filtered
        logger.info("Filtered to %d files within date range", len(files_with_dates))

    logger.info("Found %d files to import", len(files_with_dates))
    return import_batch(files_with_dates, preview=preview)


# ---------------------------------------------------------------------------
# Queue management helpers
# ---------------------------------------------------------------------------

def get_queue_status() -> Dict[str, Any]:
    """
    Get current processing queue status.

    Returns:
        Dict with queue counts by status and priority.
    """
    with get_db(readonly=True) as conn:
        total = conn.execute(
            "SELECT COUNT(*) as n FROM processing_queue"
        ).fetchone()['n']

        pending = conn.execute(
            "SELECT COUNT(*) as n FROM processing_queue WHERE status = 'pending'"
        ).fetchone()['n']

        processing = conn.execute(
            "SELECT COUNT(*) as n FROM processing_queue WHERE status = 'processing'"
        ).fetchone()['n']

        completed = conn.execute(
            "SELECT COUNT(*) as n FROM processing_queue WHERE status = 'completed'"
        ).fetchone()['n']

        failed = conn.execute(
            "SELECT COUNT(*) as n FROM processing_queue WHERE status = 'failed'"
        ).fetchone()['n']

        by_priority = {}
        rows = conn.execute(
            "SELECT priority, COUNT(*) as n FROM processing_queue "
            "WHERE status = 'pending' GROUP BY priority"
        ).fetchall()
        for r in rows:
            by_priority[r['priority']] = r['n']

    return {
        'total': total,
        'pending': pending,
        'processing': processing,
        'completed': completed,
        'failed': failed,
        'by_priority': by_priority,
    }


def list_queue_items(
    status: Optional[str] = None,
    limit: int = 20,
) -> List[Dict[str, Any]]:
    """
    List items in the processing queue.

    Args:
        status: Filter by status (pending, processing, completed, failed).
        limit: Maximum number of items to return.

    Returns:
        List of queue item dicts.
    """
    with get_db(readonly=True) as conn:
        if status:
            rows = conn.execute(
                "SELECT pq.*, s.drawing_number, s.discipline, p.number as project_number "
                "FROM processing_queue pq "
                "LEFT JOIN sheets s ON pq.sheet_id = s.id "
                "LEFT JOIN projects p ON pq.project_id = p.id "
                "WHERE pq.status = ? "
                "ORDER BY "
                "  CASE pq.priority WHEN 'high' THEN 0 WHEN 'normal' THEN 1 ELSE 2 END, "
                "  pq.created_at "
                "LIMIT ?",
                (status, limit),
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT pq.*, s.drawing_number, s.discipline, p.number as project_number "
                "FROM processing_queue pq "
                "LEFT JOIN sheets s ON pq.sheet_id = s.id "
                "LEFT JOIN projects p ON pq.project_id = p.id "
                "ORDER BY "
                "  CASE pq.priority WHEN 'high' THEN 0 WHEN 'normal' THEN 1 ELSE 2 END, "
                "  pq.created_at "
                "LIMIT ?",
                (limit,),
            ).fetchall()
    return [dict(r) for r in rows]


def add_to_queue(
    sheet_id: int,
    project_id: int,
    task: str,
    priority: str = "normal",
) -> int:
    """
    Add an item to the processing queue.

    Args:
        sheet_id: Sheet database ID.
        project_id: Project database ID.
        task: Task description (e.g. 'extract', 'review').
        priority: 'high', 'normal', or 'low'.

    Returns:
        Queue item ID.
    """
    with get_db() as conn:
        cursor = conn.execute(
            "INSERT INTO processing_queue (sheet_id, project_id, task, priority) "
            "VALUES (?, ?, ?, ?)",
            (sheet_id, project_id, task, priority),
        )
        conn.commit()
        return cursor.lastrowid
