"""
Core import engine — file parsing, column mapping, action planning, execution.

Module-agnostic. Delegates to ImportSpec callbacks for match/categorize/execute.
"""

import csv
import hashlib
import io
import json
import sqlite3
import uuid
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

from qms.imports.specs import ActionItem, ActionPlan, ColumnDef, ImportSpec

# ---------------------------------------------------------------------------
# In-memory file cache (keyed by session_id, cleared on execute/cancel)
# ---------------------------------------------------------------------------
_FILE_CACHE: Dict[str, Tuple[List[str], List[List[str]]]] = {}


def _cache_file(session_id: str, headers: List[str], rows: List[List[str]]):
    _FILE_CACHE[session_id] = (headers, rows)


def _get_cached(session_id: str) -> Optional[Tuple[List[str], List[List[str]]]]:
    return _FILE_CACHE.get(session_id)


def _clear_cache(session_id: str):
    _FILE_CACHE.pop(session_id, None)


# ---------------------------------------------------------------------------
# File parsing
# ---------------------------------------------------------------------------

def parse_file(file_bytes: bytes, filename: str) -> Tuple[List[str], List[List[str]]]:
    """Parse CSV or XLSX bytes into (headers, rows).

    Returns:
        Tuple of (header_list, row_list) where each row is a list of strings.
    """
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    if ext == "csv":
        return _parse_csv(file_bytes)
    elif ext in ("xlsx", "xls"):
        return _parse_xlsx(file_bytes)
    else:
        raise ValueError(f"Unsupported file type: .{ext} (expected .csv or .xlsx)")


def _parse_csv(data: bytes) -> Tuple[List[str], List[List[str]]]:
    # Try UTF-8 first, fall back to latin-1
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = data.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    else:
        text = data.decode("latin-1")

    reader = csv.reader(io.StringIO(text))
    rows_raw = list(reader)

    if not rows_raw:
        raise ValueError("CSV file is empty")

    headers = [h.strip() for h in rows_raw[0]]
    rows = []
    for r in rows_raw[1:]:
        if any(cell.strip() for cell in r):  # skip blank rows
            # Pad short rows
            padded = r + [""] * max(0, len(headers) - len(r))
            rows.append([c.strip() for c in padded[: len(headers)]])

    return headers, rows


def _parse_xlsx(data: bytes) -> Tuple[List[str], List[List[str]]]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("openpyxl is required for XLSX import: pip install openpyxl")

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active

    all_rows = []
    for row in ws.iter_rows(values_only=True):
        all_rows.append([str(c).strip() if c is not None else "" for c in row])

    wb.close()

    if not all_rows:
        raise ValueError("XLSX file is empty")

    headers = all_rows[0]
    rows = [r for r in all_rows[1:] if any(c for c in r)]
    return headers, rows


# ---------------------------------------------------------------------------
# Column mapping
# ---------------------------------------------------------------------------

def auto_map_columns(
    headers: List[str], spec: ImportSpec
) -> Dict[str, str]:
    """Auto-map file column indices to spec field names.

    Returns:
        Dict mapping str(column_index) -> field_name
    """
    mapping: Dict[str, str] = {}
    used_fields: set = set()

    for idx, header in enumerate(headers):
        h = header.lower().strip()
        if not h:
            continue

        for col_def in spec.columns:
            if col_def.name in used_fields:
                continue
            if h in col_def.all_names():
                mapping[str(idx)] = col_def.name
                used_fields.add(col_def.name)
                break

    return mapping


def validate_mapping(
    mapping: Dict[str, str], spec: ImportSpec
) -> List[str]:
    """Validate that all required columns are mapped.

    Returns:
        List of error messages (empty = valid).
    """
    mapped_fields = set(mapping.values())
    errors = []
    for col in spec.required_columns:
        if col.name not in mapped_fields:
            errors.append(f"Required field '{col.label}' is not mapped")
    return errors


# ---------------------------------------------------------------------------
# Row transformation
# ---------------------------------------------------------------------------

def transform_rows(
    rows: List[List[str]],
    mapping: Dict[str, str],
    spec: ImportSpec,
    conn: sqlite3.Connection,
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    """Transform raw row data using the column mapping and type coercion.

    Returns:
        Tuple of (records, parse_errors) where records is a list of dicts
        keyed by field name, and parse_errors logs any rows with issues.
    """
    records: List[Dict[str, Any]] = []
    errors: List[Dict[str, Any]] = []
    col_map = spec.column_map

    # Pre-fetch FK lookup tables
    fk_cache: Dict[str, Dict[str, int]] = {}
    for col_def in spec.columns:
        if col_def.type == "fk_lookup" and col_def.fk_table and col_def.fk_display:
            key = f"{col_def.fk_table}.{col_def.fk_display}"
            if key not in fk_cache:
                try:
                    fk_id_col = col_def.fk_id or "id"
                    rows_fk = conn.execute(
                        f"SELECT {fk_id_col}, {col_def.fk_display} FROM {col_def.fk_table}"
                    ).fetchall()
                    fk_cache[key] = {
                        str(r[col_def.fk_display]).lower(): r[fk_id_col]
                        for r in rows_fk
                    }
                except Exception:
                    fk_cache[key] = {}

    for row_idx, row in enumerate(rows):
        record: Dict[str, Any] = {"_row_index": row_idx}
        row_errors = []

        for col_idx_str, field_name in mapping.items():
            col_idx = int(col_idx_str)
            if col_idx >= len(row):
                continue

            raw = row[col_idx]
            col_def = col_map.get(field_name)
            if not col_def:
                record[field_name] = raw
                continue

            value = _coerce_value(raw, col_def, fk_cache)
            if value is _COERCE_ERROR:
                row_errors.append(f"Field '{col_def.label}': cannot parse '{raw}' as {col_def.type}")
                record[field_name] = raw  # keep raw for review
            else:
                record[field_name] = value

        if row_errors:
            errors.append({"row_index": row_idx, "errors": row_errors, "data": record})

        records.append(record)

    return records, errors


class _CoerceError:
    """Sentinel for failed type coercion."""
    pass


_COERCE_ERROR = _CoerceError()


def _coerce_value(raw: str, col_def: ColumnDef, fk_cache: Dict) -> Any:
    """Coerce a raw string value to the column's declared type."""
    if not raw and not col_def.required:
        return None

    t = col_def.type

    if t == "text":
        return raw

    if t == "int":
        try:
            return int(float(raw))
        except (ValueError, TypeError):
            return _COERCE_ERROR

    if t == "float":
        try:
            return float(raw)
        except (ValueError, TypeError):
            return _COERCE_ERROR

    if t == "bool":
        return raw.lower() in ("1", "true", "yes", "y", "x")

    if t == "date":
        # Accept common date formats, normalize to YYYY-MM-DD
        return _parse_date(raw)

    if t == "fk_lookup":
        if not raw:
            return None
        key = f"{col_def.fk_table}.{col_def.fk_display}"
        lookup = fk_cache.get(key, {})
        resolved = lookup.get(raw.lower())
        if resolved is not None:
            return resolved
        # Try numeric pass-through (already an ID)
        try:
            return int(raw)
        except (ValueError, TypeError):
            return _COERCE_ERROR

    return raw


def _parse_date(raw: str) -> Any:
    """Try common date formats and return YYYY-MM-DD or sentinel."""
    from datetime import datetime as dt

    formats = [
        "%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y",
        "%m/%d/%y", "%m-%d-%y",
        "%d/%m/%Y", "%d-%m-%Y",
        "%Y/%m/%d",
        "%B %d, %Y", "%b %d, %Y",
    ]
    for fmt in formats:
        try:
            return dt.strptime(raw.strip(), fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return _COERCE_ERROR


# ---------------------------------------------------------------------------
# Action plan generation
# ---------------------------------------------------------------------------

def generate_action_plan(
    conn: sqlite3.Connection,
    records: List[Dict[str, Any]],
    spec: ImportSpec,
    session_id: str,
    detect_missing: bool = False,
) -> ActionPlan:
    """Run match + categorize for each record, build action plan."""
    plan = ActionPlan(session_id=session_id)

    for record in records:
        row_idx = record.get("_row_index", 0)

        # Match against existing data
        existing, match_method = None, None
        if spec.match_fn:
            existing, match_method = spec.match_fn(conn, record)

        # Categorize
        if spec.categorize_fn:
            item = spec.categorize_fn(record, existing, match_method, row_idx)
        else:
            # Default: insert if no match, skip if match
            if existing:
                item = ActionItem(
                    row_index=row_idx,
                    action_type="skip",
                    record_data=record,
                    existing_data=existing,
                    match_method=match_method,
                    reason="Already exists",
                )
            else:
                item = ActionItem(
                    row_index=row_idx,
                    action_type="insert",
                    record_data=record,
                    reason="New record",
                )

        plan.items.append(item)

    # Optionally detect records in DB but not in import
    if detect_missing and spec.detect_missing_fn and len(records) >= 10:
        import_identifiers = records  # Pass full records; spec decides how to identify
        missing_items = spec.detect_missing_fn(conn, import_identifiers)
        plan.items.extend(missing_items)

    return plan


# ---------------------------------------------------------------------------
# Persist plan to DB
# ---------------------------------------------------------------------------

def save_action_plan(conn: sqlite3.Connection, plan: ActionPlan):
    """Persist all action items to import_actions table."""
    for item in plan.items:
        row = item.to_db_row(plan.session_id)
        conn.execute(
            """INSERT INTO import_actions
               (session_id, row_index, action_type, record_data, existing_data,
                match_method, changes, reason, approved, executed, execution_error)
               VALUES (:session_id, :row_index, :action_type, :record_data,
                       :existing_data, :match_method, :changes, :reason,
                       :approved, :executed, :execution_error)""",
            row,
        )
    conn.commit()


# ---------------------------------------------------------------------------
# Execution
# ---------------------------------------------------------------------------

def execute_approved_actions(
    conn: sqlite3.Connection,
    plan: ActionPlan,
    spec: ImportSpec,
    executed_by: str,
) -> Dict[str, int]:
    """Execute all approved actions. Returns summary counts."""
    counts: Dict[str, int] = {"executed": 0, "skipped": 0, "errors": 0}

    for item in plan.items:
        if not item.approved:
            counts["skipped"] += 1
            continue

        if item.action_type == "skip":
            counts["skipped"] += 1
            continue

        try:
            if spec.execute_fn:
                spec.execute_fn(conn, item, executed_by)
            item.executed = True
            counts["executed"] += 1
        except Exception as exc:
            item.execution_error = str(exc)
            counts["errors"] += 1

    conn.commit()
    return counts


# ---------------------------------------------------------------------------
# Session management
# ---------------------------------------------------------------------------

def create_import_session(
    conn: sqlite3.Connection,
    user_id: str,
    module: str,
    spec_name: str,
    filename: str,
    total_rows: int,
    file_hash: Optional[str] = None,
) -> str:
    """Create a new import session. Returns session_id."""
    session_id = str(uuid.uuid4())
    conn.execute(
        """INSERT INTO import_sessions
           (id, user_id, module, spec_name, filename, file_hash, total_rows, status)
           VALUES (?, ?, ?, ?, ?, ?, ?, 'mapping')""",
        (session_id, user_id, module, spec_name, filename, file_hash, total_rows),
    )
    conn.commit()
    return session_id


def update_session_status(
    conn: sqlite3.Connection,
    session_id: str,
    status: str,
    column_mapping: Optional[Dict] = None,
    result_summary: Optional[Dict] = None,
    error_message: Optional[str] = None,
):
    """Update session status and optional fields."""
    sets = ["status = ?"]
    vals: list = [status]

    if column_mapping is not None:
        sets.append("column_mapping = ?")
        vals.append(json.dumps(column_mapping))

    if result_summary is not None:
        sets.append("result_summary = ?")
        vals.append(json.dumps(result_summary))

    if error_message is not None:
        sets.append("error_message = ?")
        vals.append(error_message)

    if status in ("completed", "cancelled", "error"):
        sets.append("completed_at = ?")
        vals.append(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

    vals.append(session_id)
    conn.execute(
        f"UPDATE import_sessions SET {', '.join(sets)} WHERE id = ?",
        vals,
    )
    conn.commit()


def get_import_session(conn: sqlite3.Connection, session_id: str) -> Optional[Dict]:
    """Get a single import session."""
    row = conn.execute(
        "SELECT * FROM import_sessions WHERE id = ?", (session_id,)
    ).fetchone()
    return dict(row) if row else None


def get_user_import_history(
    conn: sqlite3.Connection, user_id: str, module: str, limit: int = 10
) -> List[Dict]:
    """Recent import sessions for a user + module."""
    rows = conn.execute(
        "SELECT id, filename, total_rows, status, result_summary, "
        "created_at, completed_at, error_message "
        "FROM import_sessions WHERE user_id = ? AND module = ? "
        "ORDER BY created_at DESC LIMIT ?",
        (user_id, module, limit),
    ).fetchall()
    return [dict(r) for r in rows]


def file_hash(data: bytes) -> str:
    """SHA-256 hex digest of file content."""
    return hashlib.sha256(data).hexdigest()
