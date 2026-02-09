"""
Pipeline Extraction Processor

Core SIS extraction engine: reads raw SIS sheets from Excel workbooks,
parses jobsite and personnel data, handles employee import integration,
and manages welder continuity events.

This is the main processing module that replaces sis_process_and_import.py.
It reads the raw "SIS" worksheet directly (no pre-processing needed),
parses job numbers/addresses/personnel, supports the jobs/departments
schema with customer hierarchy, and provides optional processed Excel output.
"""

import re
import sqlite3
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from qms.core import get_config, get_db, get_logger

from .common import (
    extract_date_from_filename,
    extract_department_number,
    extract_project_number,
    extract_suffix,
    load_departments_from_config,
    normalize_job_numbers,
    parse_address,
    strip_city_state,
)

logger = get_logger("qms.pipeline.processor")


# ---------------------------------------------------------------------------
# Data structures
# ---------------------------------------------------------------------------

class JobRecord:
    """Represents a job (scope of work within a department at a project)."""

    def __init__(
        self,
        job_number: str,
        scope_name: str,
        pm: str,
        street: str,
        city: str,
        state: str,
        zip_code: str,
    ):
        self.job_number = job_number
        self.scope_name = scope_name
        self.pm = pm
        self.street = street
        self.city = city
        self.state = state
        self.zip = zip_code


class EmployeeRecord:
    """Represents a parsed employee/field personnel entry."""

    def __init__(
        self,
        empl_number: str,
        last_name: str,
        first_name: str,
        job_number: str,
        phone: str,
        designation: str,
    ):
        self.employee_number = empl_number
        self.last_name = last_name
        self.first_name = first_name
        self.job_number = job_number
        self.phone = phone
        self.designation = designation


# ---------------------------------------------------------------------------
# SIS sheet parsing functions (from Office Script logic)
# ---------------------------------------------------------------------------

def parse_employee(
    empl_num: str,
    full_name: str,
    phone: str,
    job_num: str,
) -> Optional[EmployeeRecord]:
    """
    Parse employee record from name string.

    Handles:
        - Asterisks for superintendent designation (* = Super, ** = Senior, *** = General)
        - Quote removal
        - Last, First name splitting

    Args:
        empl_num: Employee number string.
        full_name: Full name, typically 'Last, First' format.
        phone: Phone number string.
        job_num: Assigned job number.

    Returns:
        EmployeeRecord or None if name is empty/invalid.
    """
    if not full_name or str(full_name).strip() == "":
        return None

    name = str(full_name).strip()
    designation = ""

    # Extract asterisk designation
    asterisk_match = re.search(r'(\*+)\s*$', name)
    if asterisk_match:
        count = len(asterisk_match.group(1))
        if count == 1:
            designation = "Superintendent"
        elif count == 2:
            designation = "Senior Superintendent"
        elif count >= 3:
            designation = "General Superintendent"
        name = re.sub(r'\*+\s*$', '', name).strip()

    # Remove quotes
    name = name.replace('"', '').strip()

    # Split Last, First
    last_name = name
    first_name = ""

    if ',' in name:
        parts = name.split(',', 1)
        last_name = parts[0].strip()
        first_name = parts[1].strip() if len(parts) > 1 else ""

    return EmployeeRecord(
        empl_number=str(empl_num).strip() if empl_num else "",
        last_name=last_name,
        first_name=first_name,
        job_number=job_num,
        phone=str(phone).strip() if phone else "",
        designation=designation,
    )


def parse_sis_sheet(wb) -> Tuple[List[JobRecord], List[EmployeeRecord]]:
    """
    Parse the raw "SIS" sheet into jobsite and employee records.

    Replicates the Office Script parsing logic: identifies job blocks by
    job number patterns, extracts PM and address rows, and collects
    employee entries within each block.

    Args:
        wb: openpyxl Workbook with an "SIS" sheet.

    Returns:
        Tuple of (list of JobRecords, list of EmployeeRecords).

    Raises:
        ValueError: If no "SIS" sheet found, or sheet is empty/malformed.
    """
    if "SIS" not in wb.sheetnames:
        raise ValueError("No worksheet named 'SIS' found in workbook")

    ws = wb["SIS"]

    values = []
    for row in ws.iter_rows(values_only=True):
        values.append(list(row))

    if len(values) < 3:
        raise ValueError("SIS sheet appears to be empty or malformed")

    # Patterns
    job_number_pattern = re.compile(r'^\d{4,5}-\d{3}')
    address_pattern = re.compile(r'^Address:', re.IGNORECASE)
    pm_pattern = re.compile(r'^PM:\s*', re.IGNORECASE)

    jobsites: List[JobRecord] = []
    employees: List[EmployeeRecord] = []
    seen_job_numbers: set = set()

    # Find first job row (skip headers)
    header_rows = 2
    first_job_row = -1

    for i in range(header_rows, len(values)):
        col_a = str(values[i][0]).strip() if values[i][0] else ""
        if job_number_pattern.match(col_a):
            first_job_row = i
            break

    if first_job_row == -1:
        raise ValueError("No job numbers found in SIS sheet")

    # Parse unassigned employees (before first job)
    for i in range(header_rows, first_job_row):
        col_b = str(values[i][1]).strip() if len(values[i]) > 1 and values[i][1] else ""
        col_c = str(values[i][2]).strip() if len(values[i]) > 2 and values[i][2] else ""
        col_d = str(values[i][3]).strip() if len(values[i]) > 3 and values[i][3] else ""

        if col_c:
            emp = parse_employee(col_b, col_c, col_d, "")
            if emp:
                employees.append(emp)

    # Parse jobsite sections
    project_blocks: List[Dict[str, Any]] = []
    current_block: Optional[Dict[str, Any]] = None

    for i in range(first_job_row, len(values)):
        col_a = str(values[i][0]).strip() if values[i][0] else ""
        col_b = str(values[i][1]).strip() if len(values[i]) > 1 and values[i][1] else ""
        col_c = str(values[i][2]).strip() if len(values[i]) > 2 and values[i][2] else ""
        col_d = str(values[i][3]).strip() if len(values[i]) > 3 and values[i][3] else ""

        # Start of job block
        if job_number_pattern.match(col_a):
            if current_block:
                project_blocks.append(current_block)
            current_block = {
                'job_num': col_a,
                'project_name': col_b or "",
                'pm': "",
                'employees': [],
            }

        # PM row
        elif pm_pattern.match(col_a):
            if current_block:
                current_block['pm'] = pm_pattern.sub('', col_a).strip()
                if col_c:
                    current_block['employees'].append({
                        'empl': col_b,
                        'name': col_c,
                        'phone': col_d,
                    })

        # Address row - finalize blocks
        elif address_pattern.match(col_a):
            raw_address = address_pattern.sub('', col_a).strip()
            street, city, state, zip_code = parse_address(raw_address)

            if current_block:
                project_blocks.append(current_block)
                current_block = None

            # Process accumulated blocks
            for block in project_blocks:
                clean_name = strip_city_state(block['project_name'])
                normalized_jobs = normalize_job_numbers(block['job_num'])

                for nj in normalized_jobs:
                    if nj not in seen_job_numbers:
                        seen_job_numbers.add(nj)
                        jobsites.append(JobRecord(
                            job_number=nj,
                            scope_name=clean_name,
                            pm=block['pm'],
                            street=street,
                            city=city,
                            state=state,
                            zip_code=zip_code,
                        ))

                # Add employees for this block
                for emp_data in block['employees']:
                    emp = parse_employee(
                        emp_data['empl'],
                        emp_data['name'],
                        emp_data['phone'],
                        normalized_jobs[0],
                    )
                    if emp:
                        employees.append(emp)

            project_blocks = []

        # Employee row (within a job block)
        elif col_c and current_block:
            current_block['employees'].append({
                'empl': col_b,
                'name': col_c,
                'phone': col_d,
            })

    # Handle remaining blocks without closing address
    if current_block:
        project_blocks.append(current_block)

    for block in project_blocks:
        clean_name = strip_city_state(block['project_name'])
        normalized_jobs = normalize_job_numbers(block['job_num'])

        for nj in normalized_jobs:
            if nj not in seen_job_numbers:
                seen_job_numbers.add(nj)
                jobsites.append(JobRecord(
                    job_number=nj,
                    scope_name=clean_name,
                    pm=block['pm'],
                    street="",
                    city="",
                    state="",
                    zip_code="",
                ))

        for emp_data in block['employees']:
            emp = parse_employee(
                emp_data['empl'],
                emp_data['name'],
                emp_data['phone'],
                normalized_jobs[0],
            )
            if emp:
                employees.append(emp)

    # Normalize all employee job numbers
    for emp in employees:
        if emp.job_number:
            normalized = normalize_job_numbers(emp.job_number)
            emp.job_number = normalized[0]

    return jobsites, employees


# ---------------------------------------------------------------------------
# Optional: save processed Excel output
# ---------------------------------------------------------------------------

def save_processed_output(
    jobsites: List[JobRecord],
    employees: List[EmployeeRecord],
    output_path: Path,
) -> None:
    """
    Save processed data to Excel (Jobsites + Field Personnel sheets).

    Args:
        jobsites: List of parsed JobRecord objects.
        employees: List of parsed EmployeeRecord objects.
        output_path: Destination path for the Excel file.
    """
    import openpyxl
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)

    # Create Jobsites sheet
    js_sheet = wb.create_sheet("Jobsites")
    js_sheet.append(["Job #", "Scope Name", "PM", "Address", "City", "State", "Zip"])

    for js in jobsites:
        js_sheet.append([
            js.job_number, js.scope_name, js.pm,
            js.street, js.city, js.state, js.zip,
        ])

    for cell in js_sheet[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    # Create Field Personnel sheet
    fp_sheet = wb.create_sheet("Field Personnel")
    fp_sheet.append(["EMPL #", "Last Name", "First Name", "Job #", "Phone Number", "Designation"])

    for emp in employees:
        fp_sheet.append([
            emp.employee_number, emp.last_name, emp.first_name,
            emp.job_number, emp.phone, emp.designation,
        ])

    for cell in fp_sheet[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    wb.save(str(output_path))
    logger.info("Saved processed output to: %s", output_path)


# ---------------------------------------------------------------------------
# Customer management
# ---------------------------------------------------------------------------

def find_customer_id(
    conn: sqlite3.Connection, customer_name: str
) -> Optional[int]:
    """Look up customer ID by name (case-insensitive)."""
    if not customer_name or not customer_name.strip():
        return None
    cursor = conn.execute(
        "SELECT id FROM customers WHERE LOWER(name) = LOWER(?)",
        (customer_name.strip(),)
    )
    row = cursor.fetchone()
    return row['id'] if row else None


def ensure_customer(
    conn: sqlite3.Connection, customer_name: str
) -> Optional[int]:
    """
    Ensure customer exists, create if missing.

    Args:
        conn: Database connection.
        customer_name: Customer name to look up or create.

    Returns:
        Customer database ID, or None if name is empty.
    """
    if not customer_name or not customer_name.strip():
        return None

    customer_id = find_customer_id(conn, customer_name)
    if customer_id:
        return customer_id

    cursor = conn.execute(
        "INSERT INTO customers (name, status) VALUES (?, 'active')",
        (customer_name.strip(),)
    )
    conn.commit()
    logger.info("Created new customer: %s", customer_name)
    return cursor.lastrowid


# ---------------------------------------------------------------------------
# Department management
# ---------------------------------------------------------------------------

def find_department_id(
    conn: sqlite3.Connection, dept_number: str
) -> Optional[int]:
    """Look up department ID by number."""
    cursor = conn.execute(
        "SELECT id FROM departments WHERE department_number = ?",
        (dept_number,)
    )
    row = cursor.fetchone()
    return row['id'] if row else None


def ensure_department(conn: sqlite3.Connection, dept_number: str) -> int:
    """
    Ensure department exists, create with generic name if missing.

    Args:
        conn: Database connection.
        dept_number: Department number (e.g. '650').

    Returns:
        Department database ID.
    """
    dept_id = find_department_id(conn, dept_number)
    if dept_id:
        return dept_id

    dept_name = f"Department {dept_number}"
    full_name = f"{dept_number}-{dept_name}"

    cursor = conn.execute(
        "INSERT INTO departments (department_number, name, full_name) VALUES (?, ?, ?)",
        (dept_number, dept_name, full_name)
    )
    conn.commit()
    logger.warning("Auto-created department: %s (%s)", dept_number, dept_name)
    return cursor.lastrowid


# ---------------------------------------------------------------------------
# Project management
# ---------------------------------------------------------------------------

def upsert_project(
    conn: sqlite3.Connection,
    project_number: str,
    project_name: str,
    customer_name: Optional[str] = None,
    street: Optional[str] = None,
    city: Optional[str] = None,
    state: Optional[str] = None,
    zip_code: Optional[str] = None,
    pm: Optional[str] = None,
) -> Tuple[Optional[int], bool]:
    """
    Find or create project by 5-digit number, optionally updating fields.

    Args:
        conn: Database connection.
        project_number: 5-digit project prefix.
        project_name: Human-readable project name.
        customer_name: Optional customer name for linkage.
        street: Optional street address.
        city: Optional city.
        state: Optional state code.
        zip_code: Optional zip code.
        pm: Optional project manager name.

    Returns:
        (project_id, is_new) - database ID and whether it was just created.
    """
    customer_id = None
    if customer_name:
        customer_id = ensure_customer(conn, customer_name)

    cursor = conn.execute(
        "SELECT id FROM projects WHERE number = ?", (project_number,)
    )
    row = cursor.fetchone()

    if row:
        project_id = row['id']
        updates = []
        params = []

        if customer_id is not None:
            updates.append("customer_id = ?")
            params.append(customer_id)
        if street and street.strip():
            updates.append("street = ?")
            params.append(street.strip())
        if city and city.strip():
            updates.append("city = ?")
            params.append(city.strip())
        if state and state.strip():
            updates.append("state = ?")
            params.append(state.strip())
        if zip_code and zip_code.strip():
            updates.append("zip = ?")
            params.append(zip_code.strip())
        if pm and pm.strip():
            updates.append("pm = ?")
            params.append(pm.strip())

        if updates:
            params.append(project_id)
            conn.execute(
                f"UPDATE projects SET {', '.join(updates)} WHERE id = ?",
                params
            )
            conn.commit()
            logger.debug("Updated project %s with new fields", project_number)

        return project_id, False

    cursor = conn.execute(
        """INSERT INTO projects (number, name, customer_id, street, city, state, zip, pm, status)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'active')""",
        (project_number, project_name, customer_id,
         street or '', city or '', state or '', zip_code or '', pm or '')
    )
    conn.commit()
    logger.info("Created NEW project: %s (%s)", project_number, project_name)
    return cursor.lastrowid, True


def upsert_job(
    conn: sqlite3.Connection,
    job: JobRecord,
    project_id: int,
    department_id: int,
    import_date: date,
    has_personnel: bool = True,
) -> None:
    """
    Upsert a job record.

    Parses job number components (project_number, department_number, suffix).
    Sets status based on personnel presence: 'active' if has_personnel,
    'inactive' if no one is assigned.

    Args:
        conn: Database connection.
        job: JobRecord object.
        project_id: Project database ID.
        department_id: Department database ID.
        import_date: Date for last_updated field.
        has_personnel: Whether this job has personnel assigned.
    """
    project_number = extract_project_number(job.job_number)
    department_number = extract_department_number(job.job_number)
    suffix = extract_suffix(job.job_number)

    status = 'active' if has_personnel else 'inactive'

    conn.execute("""
        INSERT INTO jobs (job_number, project_id, department_id,
                         project_number, department_number, suffix,
                         scope_name, pm, status, last_updated)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(job_number) DO UPDATE SET
            pm = CASE
                WHEN excluded.pm != '' AND excluded.pm != pm THEN excluded.pm
                ELSE pm
            END,
            scope_name = CASE
                WHEN excluded.scope_name != '' AND excluded.scope_name != scope_name THEN excluded.scope_name
                ELSE scope_name
            END,
            status = excluded.status,
            last_updated = CASE
                WHEN excluded.status = 'active' THEN excluded.last_updated
                ELSE last_updated
            END,
            updated_at = CURRENT_TIMESTAMP
    """, (
        job.job_number, project_id, department_id,
        project_number, department_number, suffix,
        job.scope_name, job.pm or '', status, import_date.isoformat()
    ))


# ---------------------------------------------------------------------------
# Welder continuity functions
# ---------------------------------------------------------------------------

def find_welder(
    conn: sqlite3.Connection, employee_number: str
) -> Optional[Dict[str, Any]]:
    """
    Find welder by employee number.

    Args:
        conn: Database connection.
        employee_number: Employee number to look up.

    Returns:
        Welder dict or None.
    """
    cursor = conn.execute(
        "SELECT id, employee_number, welder_stamp, display_name, status "
        "FROM weld_welder_registry WHERE employee_number = ?",
        (employee_number,)
    )
    row = cursor.fetchone()
    return dict(row) if row else None


def get_all_wpq_processes(
    conn: sqlite3.Connection, welder_id: int
) -> List[str]:
    """
    Get distinct WPQ process types for a welder (all statuses).

    Used for historical imports where we want to log production welds
    for ALL qualified processes, even if the WPQ is now inactive/expired.
    The database trigger only extends ACTIVE WPQs, which is correct.

    Args:
        conn: Database connection.
        welder_id: Welder registry ID.

    Returns:
        List of process type strings.
    """
    cursor = conn.execute(
        "SELECT DISTINCT process_type FROM weld_wpq WHERE welder_id = ?",
        (welder_id,)
    )
    return [row['process_type'] for row in cursor.fetchall()]


def get_wpq_process_details(
    conn: sqlite3.Connection,
    welder_id: int,
    active_only: bool = False,
) -> List[Dict[str, Any]]:
    """
    Get WPQ process details for a welder, including wpq_id for junction linking.

    When multiple WPQs exist for the same process, returns the active one
    (or most recent if none active).

    Args:
        conn: Database connection.
        welder_id: Welder registry ID.
        active_only: If True, only return active WPQs.

    Returns:
        List of dicts with keys: process_type, wpq_id, status.
    """
    status_filter = "AND status = 'active'" if active_only else ""
    cursor = conn.execute(f"""
        SELECT process_type, id AS wpq_id, status
        FROM weld_wpq
        WHERE welder_id = ?
        {status_filter}
        ORDER BY
            CASE status WHEN 'active' THEN 0 ELSE 1 END,
            current_expiration_date DESC
    """, (welder_id,))

    seen: Dict[str, Dict[str, Any]] = {}
    for row in cursor.fetchall():
        pt = row['process_type']
        if pt not in seen:
            seen[pt] = {
                'process_type': pt,
                'wpq_id': row['wpq_id'],
                'status': row['status'],
            }
    return list(seen.values())


def add_continuity_event(
    conn: sqlite3.Connection,
    welder_id: int,
    project_number: str,
    processes: List[Dict[str, Any]],
    event_date: date,
    week_ending: date,
    event_type: str = 'production_weld',
) -> Tuple[int, int]:
    """
    Insert continuity event with linked processes.

    Each process INSERT into the junction table fires the trigger independently,
    fixing the core bug where only the first process got its WPQ extended.

    Args:
        conn: Database connection.
        welder_id: Welder registry ID.
        project_number: Job/project number.
        processes: List of dicts with keys 'process_type', 'wpq_id'.
        event_date: Date of the welding activity.
        week_ending: Week ending date for batch imports.
        event_type: One of qualification_test, production_weld,
                    continuity_test, requalification_test.

    Returns:
        (event_id, processes_linked) tuple.
    """
    # Get employee UUID from welder_registry (if linked)
    employee_uuid = conn.execute(
        "SELECT employee_id FROM weld_welder_registry WHERE id = ?",
        (welder_id,)
    ).fetchone()
    welder_employee_id = employee_uuid[0] if employee_uuid and employee_uuid[0] else None

    # Upsert event
    conn.execute("""
        INSERT INTO weld_continuity_events (
            welder_id, welder_employee_id, event_type, event_date,
            week_ending, project_number, created_by
        ) VALUES (?, ?, ?, ?, ?, ?, 'sis_process_import')
        ON CONFLICT(welder_id, event_type, project_number, week_ending)
        DO UPDATE SET
            event_date = excluded.event_date,
            welder_employee_id = excluded.welder_employee_id
    """, (welder_id, welder_employee_id, event_type, event_date,
          week_ending, project_number))

    # Retrieve event_id (lastrowid unreliable after DO UPDATE)
    row = conn.execute("""
        SELECT id FROM weld_continuity_events
        WHERE welder_id = ? AND event_type = ?
          AND project_number = ? AND week_ending = ?
    """, (welder_id, event_type, project_number, week_ending)).fetchone()
    event_id = row['id']

    # Insert each process into junction table — trigger fires per INSERT
    processes_linked = 0
    for proc in processes:
        try:
            conn.execute("""
                INSERT OR IGNORE INTO weld_continuity_event_processes
                    (event_id, process_type, wpq_id)
                VALUES (?, ?, ?)
            """, (event_id, proc['process_type'], proc.get('wpq_id')))
            processes_linked += 1
        except sqlite3.IntegrityError:
            pass  # Already linked (idempotent)

    return event_id, processes_linked


# ---------------------------------------------------------------------------
# Schema bootstrapping
# ---------------------------------------------------------------------------

def ensure_schema(conn: sqlite3.Connection) -> None:
    """
    Create customers, departments, and jobs tables if they don't exist.

    Also loads department seed data from config.yaml if the departments
    table is empty.

    Args:
        conn: Database connection (writable).
    """
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS customers (
            id INTEGER PRIMARY KEY,
            name TEXT UNIQUE NOT NULL,
            contact_name TEXT,
            contact_email TEXT,
            contact_phone TEXT,
            billing_street TEXT,
            billing_city TEXT,
            billing_state TEXT,
            billing_zip TEXT,
            status TEXT DEFAULT 'active',
            notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        CREATE INDEX IF NOT EXISTS idx_customer_name ON customers(name);

        CREATE TABLE IF NOT EXISTS departments (
            id INTEGER PRIMARY KEY,
            department_number TEXT UNIQUE NOT NULL,
            name TEXT NOT NULL,
            full_name TEXT,
            manager TEXT,
            status TEXT DEFAULT 'active',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        CREATE INDEX IF NOT EXISTS idx_dept_number ON departments(department_number);

        CREATE TABLE IF NOT EXISTS jobs (
            id INTEGER PRIMARY KEY,
            job_number TEXT UNIQUE NOT NULL,
            project_id INTEGER NOT NULL REFERENCES projects(id),
            department_id INTEGER NOT NULL REFERENCES departments(id),
            project_number TEXT NOT NULL,
            department_number TEXT NOT NULL,
            suffix TEXT NOT NULL,
            scope_name TEXT NOT NULL,
            pm TEXT,
            status TEXT DEFAULT 'active',
            last_updated DATE,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        CREATE INDEX IF NOT EXISTS idx_jobs_project ON jobs(project_id);
        CREATE INDEX IF NOT EXISTS idx_jobs_department ON jobs(department_id);
        CREATE INDEX IF NOT EXISTS idx_jobs_status ON jobs(status);

        CREATE TABLE IF NOT EXISTS jobsites (
            id INTEGER PRIMARY KEY,
            job_number TEXT UNIQUE NOT NULL,
            project_id INTEGER REFERENCES projects(id),
            project_name TEXT,
            pm TEXT,
            street TEXT,
            city TEXT,
            state TEXT,
            zip TEXT,
            status TEXT DEFAULT 'active',
            last_updated DATE,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        CREATE INDEX IF NOT EXISTS idx_jobsites_project ON jobsites(project_id);
        CREATE INDEX IF NOT EXISTS idx_jobsites_status ON jobsites(status);
    """)

    # Load departments from config if table is empty
    cursor = conn.execute("SELECT COUNT(*) as count FROM departments")
    if cursor.fetchone()['count'] == 0:
        departments = load_departments_from_config()
        if departments:
            for dept in departments:
                conn.execute(
                    "INSERT INTO departments (department_number, name, full_name, manager) "
                    "VALUES (?, ?, ?, ?)",
                    (dept['number'], dept['name'],
                     dept.get('full_name', ''), dept.get('manager', ''))
                )
            conn.commit()
            logger.info("Loaded %d departments from config", len(departments))


# ---------------------------------------------------------------------------
# Employee import integration
# ---------------------------------------------------------------------------

def _import_employees(
    conn: sqlite3.Connection,
    employees: List[EmployeeRecord],
    week_ending: date,
    stats: Dict[str, Any],
) -> None:
    """
    Import employee data from parsed SIS records into the employees table.

    Delegates to the workforce module's employee management functions for
    duplicate detection and CRUD operations. Falls back gracefully if the
    workforce module is not available.

    Args:
        conn: Database connection.
        employees: Parsed EmployeeRecord objects.
        week_ending: Week ending date.
        stats: Mutable stats dict to update with employee counts.
    """
    try:
        from qms.workforce.employees import (
            find_employee_by_number,
            find_employee_by_phone,
            find_employee_by_name,
            create_employee,
            update_employee,
            generate_uuid,
        )
    except ImportError as e:
        logger.warning("Workforce module not available: %s", e)
        logger.warning("Skipping employee import - only welder continuity will be tracked")
        return

    for record in employees:
        stats['employees_processed'] = stats.get('employees_processed', 0) + 1

        if not record.employee_number:
            stats['employees_skipped'] = stats.get('employees_skipped', 0) + 1
            continue

        last_name = record.last_name
        first_name = record.first_name
        if not last_name or not first_name:
            logger.warning("Could not parse name for employee %s", record.employee_number)
            stats['employees_skipped'] = stats.get('employees_skipped', 0) + 1
            continue

        # 4-level duplicate matching
        existing = None
        match_method = None

        # Level 1: employee_number
        existing = find_employee_by_number(conn, employee_number=record.employee_number)
        if existing:
            match_method = 'employee_number'

        # Level 2: phone
        if not existing and record.phone and record.phone != 'nan':
            existing = find_employee_by_phone(conn, record.phone)
            if existing:
                match_method = 'phone'

        # Level 3: name match
        if not existing:
            candidates = find_employee_by_name(conn, last_name, first_name)
            if len(candidates) == 1:
                existing = candidates[0]
                match_method = 'name_fuzzy'
                stats.setdefault('duplicate_warnings', []).append({
                    'employee_number': record.employee_number,
                    'matched_id': existing['id'],
                    'reason': 'Fuzzy name match - verify not duplicate',
                })
            elif len(candidates) > 1:
                logger.warning("Multiple name matches for %s, %s - skipping",
                               last_name, first_name)
                continue

        # Get job_id from job_number
        job_id = None
        if record.job_number:
            job_row = conn.execute(
                "SELECT id FROM jobs WHERE job_number = ?",
                (record.job_number,)
            ).fetchone()
            if job_row:
                job_id = job_row['id']

        if existing:
            # Update existing employee
            current_data = conn.execute(
                "SELECT job_id, original_hire_date, current_hire_date FROM employees WHERE id = ?",
                (existing['id'],)
            ).fetchone()

            current_original_hire = current_data['original_hire_date'] if current_data else None

            # Check if this week_ending is earlier than stored hire dates
            if current_original_hire and str(week_ending) < current_original_hire:
                conn.execute("""
                    UPDATE employees
                    SET original_hire_date = ?,
                        current_hire_date = ?
                    WHERE id = ?
                """, (str(week_ending), str(week_ending), existing['id']))
                stats['hire_dates_corrected'] = stats.get('hire_dates_corrected', 0) + 1
                logger.info("Hire date updated for %s: was %s, now %s",
                            record.employee_number, current_original_hire, week_ending)

            current_job_id = current_data['job_id'] if current_data else None
            job_changed = (current_job_id != job_id) and job_id is not None

            update_employee(
                conn,
                existing['id'],
                phone=record.phone if record.phone else None,
                job_id=job_id,
                notes=record.designation if record.designation else None,
            )

            if job_changed:
                conn.execute("""
                    UPDATE employment_history
                    SET end_date = ?
                    WHERE employee_id = ? AND end_date IS NULL
                """, (str(week_ending), existing['id']))

                conn.execute("""
                    INSERT INTO employment_history (
                        id, employee_id, start_date, employment_type,
                        position, job_id, transition_type, reason_for_change, created_by
                    )
                    SELECT
                        ?, ?, ?,
                        CASE WHEN is_employee = 1 AND is_subcontractor = 1 THEN 'both'
                             WHEN is_employee = 1 THEN 'employee'
                             ELSE 'subcontractor' END,
                        position, ?, 'transfer', 'Job assignment change from SIS import', 'SIS-IMPORT'
                    FROM employees WHERE id = ?
                """, (generate_uuid(), existing['id'], str(week_ending), job_id, existing['id']))

            stats['employees_updated'] = stats.get('employees_updated', 0) + 1
            if job_id:
                stats['employees_job_assigned'] = stats.get('employees_job_assigned', 0) + 1

        else:
            # Create new employee
            is_employee = not (
                record.employee_number.startswith('CONTRACT') or
                'contractor' in (record.designation or '').lower()
            )
            is_subcontractor = not is_employee

            role_row = conn.execute(
                "SELECT id FROM roles WHERE role_code = 'TECH'"
            ).fetchone()
            role_id = role_row['id'] if role_row else None

            create_employee(
                conn,
                last_name=last_name,
                first_name=first_name,
                is_employee=is_employee,
                is_subcontractor=is_subcontractor,
                position='Field Personnel',
                job_id=job_id,
                role_id=role_id,
                phone=record.phone if record.phone else None,
                current_hire_date=str(week_ending),
                original_hire_date=str(week_ending),
                status='active',
                notes=record.designation if record.designation else None,
                created_by='SIS-IMPORT',
            )

            stats['employees_created'] = stats.get('employees_created', 0) + 1
            if job_id:
                stats['employees_job_assigned'] = stats.get('employees_job_assigned', 0) + 1

    conn.commit()


# ---------------------------------------------------------------------------
# Main processing entry point
# ---------------------------------------------------------------------------

def process_and_import(
    filepath: Path,
    week_override: Optional[date] = None,
    output_path: Optional[Path] = None,
    preview: bool = False,
) -> Dict[str, Any]:
    """
    Main unified processing logic.

    1. Parse SIS sheet from Excel workbook
    2. Optionally save processed Excel output
    3. Import to database:
       a. Upsert jobs/projects/departments
       b. Import employees to workforce module
       c. Create welder continuity events

    Args:
        filepath: Path to the Excel workbook with a raw "SIS" sheet.
        week_override: Override the week ending date.
        output_path: Optional path for processed Excel output.
        preview: If True, parse only without database changes.

    Returns:
        Statistics dict with counts for all import operations.
    """
    import openpyxl

    stats: Dict[str, Any] = {
        'jobsites_created': 0,
        'jobsites_updated': 0,
        'jobsites_active': 0,
        'jobsites_inactive': 0,
        'jobsites_processed': 0,
        'projects_created': 0,
        'projects_activated': 0,
        'projects_deactivated': 0,
        'personnel_processed': 0,
        'unassigned_personnel': 0,
        'employees_processed': 0,
        'employees_created': 0,
        'employees_updated': 0,
        'employees_job_assigned': 0,
        'employees_skipped': 0,
        'duplicate_warnings': [],
        'welders_matched': 0,
        'welders_active': 0,
        'welders_inactive': 0,
        'non_welders': 0,
        'continuity_events_created': 0,
        'continuity_processes_linked': 0,
        'errors': [],
    }

    # Determine week ending date
    if week_override:
        weld_date = week_override
        logger.info("Using week ending date from override: %s", weld_date)
    else:
        filename_date = extract_date_from_filename(filepath)
        if filename_date:
            weld_date = filename_date
            logger.info("Extracted week ending date from filename: %s", weld_date)
        else:
            weld_date = date.today()
            logger.info("No date in filename or override, using today: %s", weld_date)

    week_ending = weld_date

    # Step 1: Parse SIS sheet
    logger.info("Reading workbook: %s", filepath)
    wb = openpyxl.load_workbook(str(filepath), data_only=True)

    logger.info("Parsing SIS sheet...")
    jobsites, employees = parse_sis_sheet(wb)

    stats['jobsites_processed'] = len(jobsites)
    stats['personnel_processed'] = len(employees)
    stats['unassigned_personnel'] = sum(1 for e in employees if not e.job_number)

    logger.info("Parsed %d jobsites, %d personnel", len(jobsites), len(employees))

    # Step 2: Optionally save processed output
    if output_path:
        output_str = str(output_path).replace('{date}', weld_date.strftime('%Y-%m-%d'))
        output_path = Path(output_str)
        save_processed_output(jobsites, employees, output_path)
        logger.info("Saved processed Excel to: %s", output_path)

    if preview:
        logger.info("PREVIEW MODE - No database changes")
        logger.info("Would import %d jobsites and %d personnel", len(jobsites), len(employees))
        return stats

    # Step 3: Import to database
    logger.info("Importing to database...")

    # Build personnel-to-jobsite mapping
    jobsites_with_personnel: set = set()
    for emp in employees:
        if emp.job_number:
            jobsites_with_personnel.add(emp.job_number)

    with get_db() as conn:
        ensure_schema(conn)

        projects_to_update: Dict[int, bool] = {}

        # Import jobs
        for js in jobsites:
            proj_num = extract_project_number(js.job_number)
            dept_num = extract_department_number(js.job_number)

            if dept_num:
                department_id = ensure_department(conn, dept_num)
            else:
                logger.warning("Could not extract department from job number: %s", js.job_number)
                continue

            cursor = conn.execute("SELECT id FROM projects WHERE number = ?", (proj_num,))
            existing_proj = cursor.fetchone()

            project_id, is_new = upsert_project(
                conn, proj_num, js.scope_name,
                customer_name=None,
                street=js.street,
                city=js.city,
                state=js.state,
                zip_code=js.zip,
                pm=js.pm,
            )
            if is_new:
                stats['projects_created'] += 1

            has_personnel = js.job_number in jobsites_with_personnel
            if project_id not in projects_to_update:
                projects_to_update[project_id] = False
            if has_personnel:
                projects_to_update[project_id] = True

            cursor = conn.execute("SELECT id FROM jobs WHERE job_number = ?", (js.job_number,))
            existing_js = cursor.fetchone()
            if existing_js:
                stats['jobsites_updated'] += 1
            else:
                stats['jobsites_created'] += 1

            upsert_job(conn, js, project_id, department_id, weld_date, has_personnel)

            if has_personnel:
                stats['jobsites_active'] += 1
            else:
                stats['jobsites_inactive'] += 1

        # Update project statuses
        for project_id, has_active in projects_to_update.items():
            new_status = 'active' if has_active else 'inactive'

            cursor = conn.execute("SELECT status FROM projects WHERE id = ?", (project_id,))
            row = cursor.fetchone()
            old_status = row['status'] if row else None

            if old_status != new_status:
                conn.execute("""
                    UPDATE projects
                    SET status = ?,
                        updated_at = CURRENT_TIMESTAMP
                    WHERE id = ?
                """, (new_status, project_id))
                logger.info("Updated project %d status: %s -> %s",
                            project_id, old_status, new_status)

                if new_status == 'active':
                    stats['projects_activated'] += 1
                else:
                    stats['projects_deactivated'] += 1

        conn.commit()
        logger.info("Imported %d jobsites", len(jobsites))

        # Import personnel to employees table
        logger.info("Importing %d personnel to employees table...", len(employees))
        try:
            _import_employees(conn, employees, week_ending, stats)
        except Exception as e:
            logger.error("Error importing employees: %s", e, exc_info=True)
            stats['errors'].append(f"Employee import error: {str(e)}")

        # Process personnel for welder continuity
        for person in employees:
            if not person.job_number:
                continue

            welder = find_welder(conn, person.employee_number)
            if not welder:
                stats['non_welders'] += 1
                continue

            stats['welders_matched'] += 1
            if welder['status'] == 'active':
                stats['welders_active'] += 1
            else:
                stats['welders_inactive'] += 1

            proc_details = get_wpq_process_details(conn, welder['id'])

            if not proc_details:
                continue

            try:
                event_id, linked = add_continuity_event(
                    conn, welder['id'], person.job_number,
                    proc_details, weld_date, week_ending,
                )
                stats['continuity_events_created'] += 1
                stats['continuity_processes_linked'] += linked
                procs_str = ', '.join(p['process_type'] for p in proc_details)
                logger.info("[OK] %s: %s | %s | week %s",
                            welder.get('welder_stamp', person.employee_number),
                            person.job_number, procs_str, week_ending)
            except Exception as e:
                stats['errors'].append(f"{person.employee_number}: {str(e)}")
                logger.error("Error adding continuity event: %s", e)

        conn.commit()

    return stats


# ---------------------------------------------------------------------------
# Pipeline status query
# ---------------------------------------------------------------------------

def get_pipeline_status(project_number: Optional[str] = None) -> Dict[str, Any]:
    """
    Get pipeline processing status.

    If project_number is given, returns per-discipline breakdown for that project.
    Otherwise returns overall summary across all projects.

    Args:
        project_number: Optional project number to filter by.

    Returns:
        Status dict with counts and breakdowns.
    """
    with get_db(readonly=True) as conn:
        if project_number:
            proj = conn.execute(
                "SELECT id, number, name FROM projects WHERE number = ?",
                (project_number,)
            ).fetchone()
            if not proj:
                return {'error': f'Project {project_number} not found'}

            project_id = proj['id']

            total_sheets = conn.execute(
                "SELECT COUNT(*) as n FROM sheets WHERE project_id = ?",
                (project_id,)
            ).fetchone()['n']

            extracted = conn.execute(
                "SELECT COUNT(*) as n FROM sheets "
                "WHERE project_id = ? AND extracted_at IS NOT NULL",
                (project_id,)
            ).fetchone()['n']

            pending_queue = conn.execute(
                "SELECT COUNT(*) as n FROM processing_queue "
                "WHERE project_id = ? AND status = 'pending'",
                (project_id,)
            ).fetchone()['n']

            disc_rows = conn.execute(
                "SELECT name, sheet_count, processed_count "
                "FROM disciplines WHERE project_id = ? ORDER BY name",
                (project_id,)
            ).fetchall()

            return {
                'project_number': proj['number'],
                'project_name': proj['name'],
                'total_sheets': total_sheets,
                'extracted': extracted,
                'pending_queue': pending_queue,
                'disciplines': [dict(r) for r in disc_rows],
            }

        else:
            total_projects = conn.execute(
                "SELECT COUNT(*) as n FROM projects WHERE status = 'active'"
            ).fetchone()['n']

            total_sheets = conn.execute(
                "SELECT COUNT(*) as n FROM sheets"
            ).fetchone()['n']

            extracted = conn.execute(
                "SELECT COUNT(*) as n FROM sheets WHERE extracted_at IS NOT NULL"
            ).fetchone()['n']

            pending_queue = conn.execute(
                "SELECT COUNT(*) as n FROM processing_queue WHERE status = 'pending'"
            ).fetchone()['n']

            conflicts = conn.execute(
                "SELECT COUNT(*) as n FROM conflicts WHERE resolved = 0"
            ).fetchone()['n']

            return {
                'active_projects': total_projects,
                'total_sheets': total_sheets,
                'extracted': extracted,
                'pending_queue': pending_queue,
                'open_conflicts': conflicts,
            }
