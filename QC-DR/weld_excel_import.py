#!/usr/bin/env python3
"""
Welding Excel Import Script

Imports welder data and WPQ qualifications from the Excel spreadsheet
(Welding Daily Log.xlsm) into the database.

Features:
- Parses WPQ codes to extract process, P-number, position, filler metal
- Idempotent imports via row hash comparison
- Dry-run mode for preview
- Progress reporting

Usage:
    python weld_excel_import.py                    # Full import
    python weld_excel_import.py --dry-run          # Preview changes
    python weld_excel_import.py --validate         # Check Excel structure
    python weld_excel_import.py --welder A-8       # Import single welder

WPQ Code Patterns (examples from actual data):
    A53-NPS6-6G-6010-7018     -> Material(A53), Size(NPS6), Position(6G), Root(6010), Fill(7018) = SMAW
    SS-01-P8-GTAW             -> WPS(SS-01), P-Number(8), Process(GTAW)
    CS-01-P1-SMAW             -> WPS(CS-01), P-Number(1), Process(SMAW)
    A333-NPS6-6G-ER80S-8018   -> Material(A333), Size(NPS6), Position(6G), Root(ER80S/GTAW), Fill(8018/SMAW)
    1-6G-7                    -> P-Number(1), Position(6G), Thickness(7 = 3/4")

Created: 2026-02-05
"""

import sqlite3
import hashlib
import re
import sys
import argparse
from pathlib import Path
from datetime import datetime, date, timedelta
from typing import Optional, Dict, List, Tuple, Any

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

# =============================================================================
# CONFIGURATION
# =============================================================================

EXCEL_PATH = Path("D:/Quality Documents/Welding/Welding Daily Log.xlsm")
DB_PATH = Path("D:/quality.db")
MAIN_SHEET = "Welder_Info"
HEADER_ROW = 3
DATA_START_ROW = 4

# Column mapping (0-indexed from Excel columns A=0)
COLUMNS = {
    'status': 0,          # A: Status (Active/Inactive)
    'stamp': 1,           # B: Stamp Number
    'employee': 2,        # C: Employee #
    'display_name': 3,    # D: Display Name
    'last_name': 4,       # E: Last Name
    'first_name': 5,      # F: First Name
    'preferred_name': 6,  # G: Preferred Name
    'department': 7,      # H: Department
    'supervisor': 8,      # I: Supervisor
    'business_unit': 9,   # J: Business Unit
    'wpq1': 10,           # K: WPQ
    'wpq2': 11,           # L: WPQ2
    'wpq3': 12,           # M: WPQ3
    'wpq4': 13,           # N: WPQ4
    'wpq5': 14,           # O: WPQ5
    'running_total': 15,  # P: RunningTotalWeldsMade
    'total_tested': 16,   # Q: TotalWeldsTested
    'pass': 17,           # R: Pass
    'fail': 18,           # S: Fail
}

# =============================================================================
# WPQ CODE PARSER
# =============================================================================

# Material spec to P-number mapping
MATERIAL_P_NUMBERS = {
    'A53': 1, 'A106': 1, 'A333': 1,  # Carbon steel pipe
    'A516': 1,  # Carbon steel plate
    'A312': 8, 'A358': 8,  # Stainless steel pipe
    'A240': 8,  # Stainless steel plate
    'SS': 8, 'SS304': 8, 'SS316': 8,  # Stainless steel (generic)
    'CS': 1,  # Carbon steel (generic)
    'DM': None,  # Dissimilar metal (special case)
}

# Filler metal to F-number and process mapping
FILLER_METAL_INFO = {
    # SMAW electrodes (F-number 4)
    '6010': {'f_number': 3, 'process': 'SMAW', 'description': 'E6010'},
    '6011': {'f_number': 3, 'process': 'SMAW', 'description': 'E6011'},
    '7018': {'f_number': 4, 'process': 'SMAW', 'description': 'E7018'},
    '8018': {'f_number': 4, 'process': 'SMAW', 'description': 'E8018'},
    '8010': {'f_number': 3, 'process': 'SMAW', 'description': 'E8010'},
    # GTAW wire (F-number 6)
    'ER70S': {'f_number': 6, 'process': 'GTAW', 'description': 'ER70S-2/ER70S-6'},
    'ER70S2': {'f_number': 6, 'process': 'GTAW', 'description': 'ER70S-2'},
    'ER70S6': {'f_number': 6, 'process': 'GTAW', 'description': 'ER70S-6'},
    'ER80S': {'f_number': 6, 'process': 'GTAW', 'description': 'ER80S-Ni1'},
    'ER308': {'f_number': 6, 'process': 'GTAW', 'description': 'ER308L'},
    'ER309': {'f_number': 6, 'process': 'GTAW', 'description': 'ER309L'},
    'ER316': {'f_number': 6, 'process': 'GTAW', 'description': 'ER316L'},
}

# Position qualifications (what each position qualifies)
POSITION_QUALIFIES = {
    '1G': ['1G'],
    '2G': ['1G', '2G'],
    '3G': ['1G', '3G'],
    '4G': ['1G', '4G'],
    '5G': ['1G', '2G', '5G'],
    '6G': ['1G', '2G', '3G', '4G', '5G', '6G'],  # 6G qualifies all
    '6GR': ['1G', '2G', '3G', '4G', '5G', '6G', '6GR'],
    '1F': ['1F'],
    '2F': ['1F', '2F'],
    '3F': ['1F', '2F', '3F'],
    '4F': ['1F', '2F', '3F', '4F'],
    '5F': ['1F', '2F', '3F', '4F', '5F'],
}


def parse_wpq_code(code: str) -> Dict[str, Any]:
    """
    Parse a WPQ code and extract qualification details.

    Returns dict with:
        process_type: SMAW, GTAW, GMAW, etc.
        p_number: Base metal P-number
        f_number: Filler metal F-number
        positions: List of positions qualified
        thickness_min: Minimum thickness qualified (inches)
        thickness_max: Maximum thickness qualified (inches)
        diameter_min: Minimum diameter qualified (inches)
        wps_number: Referenced WPS if any
        raw_code: Original code
        parse_notes: Any notes about parsing
    """
    result = {
        'process_type': None,
        'p_number': None,
        'f_number': None,
        'positions': [],
        'thickness_min': None,
        'thickness_max': None,
        'diameter_min': None,
        'wps_number': None,
        'raw_code': code,
        'parse_notes': [],
    }

    if not code or code.strip() in ('', '0', 'None'):
        return result

    code = code.strip().upper()
    original = code

    # Normalize: Remove extra hyphens in material specs (A-53 -> A53, A-106 -> A106)
    code = re.sub(r'^(A)[-_](\d+)', r'\1\2', code)  # A-53 -> A53, A-106 -> A106, A-333 -> A333

    # Pattern 1: WPS reference style (SS-01-P8-GTAW, CS-01-P1-SMAW)
    # Format: WPS_PREFIX-NUMBER-P_NUMBER-PROCESS
    wps_match = re.match(r'^(SS|CS|DM)[-_]?(\d+)[-_]?P(\d+)[-_]?(GTAW|SMAW|GMAW|FCAW)?', code, re.I)
    if wps_match:
        prefix, wps_num, p_num, process = wps_match.groups()
        result['wps_number'] = f"{prefix}-{wps_num}"
        result['p_number'] = int(p_num)
        if process:
            result['process_type'] = process.upper()
        # Determine process from material prefix if not specified
        if not result['process_type']:
            if prefix == 'SS':
                result['process_type'] = 'GTAW'  # Stainless usually GTAW
                result['f_number'] = 6
            else:
                result['process_type'] = 'SMAW'  # Default
                result['f_number'] = 4
        return result

    # Pattern 2: Material-Size-Position-Fillers (A53-NPS6-6G-6010-7018)
    # Also handles variants like A53-NPS6-6G-6010-7018-3 (trailing suffix)
    mat_match = re.match(r'^([A-Z]+\d*)?[-_]?(NPS\d+|[\d.]+["\']?)?[-_]?(\d+G[R]?)?[-_]?([\dA-Z]+)?[-_]?([\dA-Z]+)?(?:[-_][\dA-Za-z]+)?', code)
    if mat_match:
        material, size, position, filler1, filler2 = mat_match.groups()

        # Material -> P-number
        if material:
            material_clean = re.sub(r'[-_]', '', material)
            for mat_key, p_num in MATERIAL_P_NUMBERS.items():
                if material_clean.startswith(mat_key):
                    result['p_number'] = p_num
                    break

        # Size -> diameter qualified
        if size:
            size_match = re.search(r'NPS(\d+)|(\d+\.?\d*)', size)
            if size_match:
                nps = size_match.group(1) or size_match.group(2)
                try:
                    result['diameter_min'] = float(nps)
                except ValueError:
                    pass

        # Position
        if position:
            pos = position.upper()
            result['positions'] = POSITION_QUALIFIES.get(pos, [pos])

        # Filler metals -> process and F-number
        processes = []
        for filler in [filler1, filler2]:
            if filler:
                filler_clean = filler.upper()
                for fm_key, fm_info in FILLER_METAL_INFO.items():
                    if filler_clean.startswith(fm_key) or filler_clean == fm_key:
                        processes.append(fm_info['process'])
                        if result['f_number'] is None or fm_info['f_number'] < result['f_number']:
                            result['f_number'] = fm_info['f_number']
                        break

        # Determine process (combo or single)
        if processes:
            unique_processes = list(set(processes))
            if len(unique_processes) == 1:
                result['process_type'] = unique_processes[0]
            elif 'GTAW' in unique_processes and 'SMAW' in unique_processes:
                result['process_type'] = 'GTAW/SMAW'  # Combo procedure
            else:
                result['process_type'] = '/'.join(unique_processes)

    # Pattern 3: Short code (1-6G-7 = P1, 6G position, thickness group 7)
    short_match = re.match(r'^(\d+)[-_](\d+G[R]?)[-_](\d+[A]?)$', code)
    if short_match:
        p_num, position, thickness_code = short_match.groups()
        result['p_number'] = int(p_num)
        result['positions'] = POSITION_QUALIFIES.get(position.upper(), [position.upper()])
        result['process_type'] = 'SMAW'  # Default for short codes
        result['f_number'] = 4  # Default

        # Thickness code to actual thickness (rough mapping)
        # These are typical ranges used in the industry
        thickness_map = {
            '7': (0.75, 999),    # 3/4" and up
            '8': (0.75, 999),    # 3/4" and up
            '8A': (0.75, 999),   # 3/4" and up with A suffix
        }
        if thickness_code in thickness_map:
            result['thickness_min'], result['thickness_max'] = thickness_map[thickness_code]

    # If still no process type, check for process keywords in code
    if not result['process_type']:
        code_upper = original.upper()
        if 'GTAW' in code_upper:
            result['process_type'] = 'GTAW'
            result['f_number'] = result['f_number'] or 6
        elif 'SMAW' in code_upper:
            result['process_type'] = 'SMAW'
            result['f_number'] = result['f_number'] or 4
        elif 'GMAW' in code_upper:
            result['process_type'] = 'GMAW'
            result['f_number'] = result['f_number'] or 6
        elif 'FCAW' in code_upper:
            result['process_type'] = 'FCAW'
            result['f_number'] = result['f_number'] or 6

    # Pattern 4: Handle incomplete codes with just material (A333, A106)
    if not result['process_type'] and result['p_number']:
        # Has material P-number but no process - assume SMAW as default
        result['process_type'] = 'SMAW'
        result['f_number'] = 4
        result['parse_notes'].append(f"Incomplete code, assumed SMAW: {original}")

    # Pattern 5: SWPS references (SWPS 6010-7018 AWS D1.1)
    swps_match = re.search(r'SWPS|AWS[-\s]*(D1|B2)', original)
    if swps_match and not result['process_type']:
        result['process_type'] = 'SMAW'
        result['f_number'] = 4
        result['wps_number'] = original  # Store full reference
        return result

    # Default fallback
    if not result['process_type']:
        result['process_type'] = 'UNKNOWN'
        result['parse_notes'].append(f"Could not determine process from code: {original}")

    return result


def compute_row_hash(row: tuple) -> str:
    """Compute hash of row data for idempotent imports."""
    # Use columns that define the welder's qualifications
    key_cols = [str(row[i]) if i < len(row) and row[i] is not None else ''
                for i in range(COLUMNS['status'], COLUMNS['fail'] + 1)]
    data = '|'.join(key_cols)
    return hashlib.md5(data.encode()).hexdigest()


# =============================================================================
# DATABASE OPERATIONS
# =============================================================================

def get_db_connection() -> sqlite3.Connection:
    """Get database connection."""
    conn = sqlite3.connect(str(DB_PATH))
    conn.execute("PRAGMA foreign_keys = ON")
    conn.row_factory = sqlite3.Row
    return conn


def find_welder_by_stamp(conn: sqlite3.Connection, stamp: str) -> Optional[Dict]:
    """Find welder by stamp number."""
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM weld_welder_registry WHERE welder_stamp = ?", (stamp,))
    row = cursor.fetchone()
    return dict(row) if row else None


def find_welder_by_employee(conn: sqlite3.Connection, emp_num: str) -> Optional[Dict]:
    """Find welder by employee number."""
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM weld_welder_registry WHERE employee_number = ?", (emp_num,))
    row = cursor.fetchone()
    return dict(row) if row else None


def upsert_welder(conn: sqlite3.Connection, welder_data: Dict, row_hash: str) -> int:
    """Insert or update welder record. Returns welder ID."""
    cursor = conn.cursor()

    # Check if exists by stamp first, then by employee number
    existing = None
    if welder_data.get('welder_stamp'):
        existing = find_welder_by_stamp(conn, welder_data['welder_stamp'])
    if not existing and welder_data.get('employee_number'):
        existing = find_welder_by_employee(conn, welder_data['employee_number'])

    if existing:
        # Update if hash changed
        if existing.get('excel_row_hash') != row_hash:
            cursor.execute("""
                UPDATE weld_welder_registry SET
                    last_name = ?,
                    first_name = ?,
                    preferred_name = ?,
                    display_name = ?,
                    department = ?,
                    supervisor = ?,
                    business_unit = ?,
                    status = ?,
                    running_total_welds = ?,
                    total_welds_tested = ?,
                    welds_passed = ?,
                    welds_failed = ?,
                    excel_row_hash = ?,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
            """, (
                welder_data.get('last_name'),
                welder_data.get('first_name'),
                welder_data.get('preferred_name'),
                welder_data.get('display_name'),
                welder_data.get('department'),
                welder_data.get('supervisor'),
                welder_data.get('business_unit'),
                welder_data.get('status', 'active'),
                welder_data.get('running_total_welds', 0),
                welder_data.get('total_welds_tested', 0),
                welder_data.get('welds_passed', 0),
                welder_data.get('welds_failed', 0),
                row_hash,
                existing['id']
            ))
            conn.commit()
        return existing['id']

    # Insert new welder
    # Ensure employee_number has a value (use stamp if no employee number)
    emp_num = welder_data.get('employee_number') or welder_data.get('welder_stamp') or f"UNK-{row_hash[:8]}"

    cursor.execute("""
        INSERT INTO weld_welder_registry (
            employee_number, welder_stamp, last_name, first_name,
            preferred_name, display_name, department, supervisor,
            business_unit, status, running_total_welds, total_welds_tested,
            welds_passed, welds_failed, excel_row_hash
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        emp_num,
        welder_data.get('welder_stamp'),
        welder_data.get('last_name'),
        welder_data.get('first_name'),
        welder_data.get('preferred_name'),
        welder_data.get('display_name'),
        welder_data.get('department'),
        welder_data.get('supervisor'),
        welder_data.get('business_unit'),
        welder_data.get('status', 'active'),
        welder_data.get('running_total_welds', 0),
        welder_data.get('total_welds_tested', 0),
        welder_data.get('welds_passed', 0),
        welder_data.get('welds_failed', 0),
        row_hash
    ))
    conn.commit()
    return cursor.lastrowid


def upsert_wpq(conn: sqlite3.Connection, welder_id: int, wpq_code: str, parsed: Dict) -> Optional[int]:
    """Insert or update WPQ record. Returns WPQ ID or None if no valid data."""
    if not parsed.get('process_type') or parsed['process_type'] == 'UNKNOWN':
        return None

    cursor = conn.cursor()

    # Generate WPQ number from welder stamp + code hash
    cursor.execute("SELECT welder_stamp FROM weld_welder_registry WHERE id = ?", (welder_id,))
    welder = cursor.fetchone()
    stamp = welder['welder_stamp'] if welder else 'UNK'

    # Create unique WPQ number
    wpq_number = f"{stamp}-{wpq_code[:20]}"

    # Check for existing
    cursor.execute("SELECT id FROM weld_wpq WHERE wpq_number = ?", (wpq_number,))
    existing = cursor.fetchone()

    # Calculate expiration: 6 months from today (import baseline)
    initial_expiration = date.today() + timedelta(days=180)

    if existing:
        # Update
        cursor.execute("""
            UPDATE weld_wpq SET
                wps_number = COALESCE(?, wps_number),
                process_type = ?,
                p_number_base = COALESCE(?, p_number_base),
                f_number = COALESCE(?, f_number),
                groove_positions_qualified = COALESCE(?, groove_positions_qualified),
                thickness_qualified_min = COALESCE(?, thickness_qualified_min),
                thickness_qualified_max = COALESCE(?, thickness_qualified_max),
                diameter_qualified_min = COALESCE(?, diameter_qualified_min),
                updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
        """, (
            parsed.get('wps_number'),
            parsed['process_type'],
            parsed.get('p_number'),
            parsed.get('f_number'),
            ', '.join(parsed.get('positions', [])) if parsed.get('positions') else None,
            parsed.get('thickness_min'),
            parsed.get('thickness_max'),
            parsed.get('diameter_min'),
            existing['id']
        ))
        conn.commit()
        return existing['id']

    # Insert new
    cursor.execute("""
        INSERT INTO weld_wpq (
            wpq_number, welder_id, welder_stamp, wps_number,
            process_type, p_number_base, f_number,
            groove_positions_qualified, thickness_qualified_min,
            thickness_qualified_max, diameter_qualified_min,
            initial_expiration_date, current_expiration_date,
            status, notes
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'active', ?)
    """, (
        wpq_number,
        welder_id,
        stamp,
        parsed.get('wps_number'),
        parsed['process_type'],
        parsed.get('p_number'),
        parsed.get('f_number'),
        ', '.join(parsed.get('positions', [])) if parsed.get('positions') else None,
        parsed.get('thickness_min'),
        parsed.get('thickness_max'),
        parsed.get('diameter_min'),
        initial_expiration,
        initial_expiration,
        f"Imported from Excel code: {wpq_code}"
    ))
    conn.commit()
    return cursor.lastrowid


# =============================================================================
# IMPORT FUNCTIONS
# =============================================================================

def validate_excel() -> bool:
    """Validate Excel file structure."""
    print(f"Validating: {EXCEL_PATH}")

    if not EXCEL_PATH.exists():
        print(f"ERROR: Excel file not found: {EXCEL_PATH}")
        return False

    try:
        wb = openpyxl.load_workbook(str(EXCEL_PATH), data_only=True)
    except Exception as e:
        print(f"ERROR: Cannot open Excel file: {e}")
        return False

    if MAIN_SHEET not in wb.sheetnames:
        print(f"ERROR: Sheet '{MAIN_SHEET}' not found. Available: {wb.sheetnames}")
        return False

    ws = wb[MAIN_SHEET]

    # Check headers
    expected_headers = ['Status', 'Stamp Number', 'Employee #', 'Display Name',
                        'Last Name', 'First Name', 'Preferred Name']
    headers = [ws.cell(row=HEADER_ROW, column=i+1).value for i in range(7)]

    print(f"Headers found: {headers}")

    for i, (expected, actual) in enumerate(zip(expected_headers, headers)):
        if expected.lower() != (actual or '').lower():
            print(f"WARNING: Column {i+1} header mismatch. Expected '{expected}', got '{actual}'")

    # Count data rows
    data_count = 0
    for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        if row[0]:  # Has status
            data_count += 1

    print(f"Data rows found: {data_count}")
    print("Validation PASSED")
    return True


def import_from_excel(dry_run: bool = False, single_welder: str = None) -> Dict:
    """
    Import welders and WPQs from Excel.

    Returns dict with import statistics.
    """
    stats = {
        'welders_processed': 0,
        'welders_created': 0,
        'welders_updated': 0,
        'welders_skipped': 0,
        'wpqs_created': 0,
        'wpqs_updated': 0,
        'wpq_parse_errors': 0,
        'errors': [],
    }

    print(f"Loading: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(str(EXCEL_PATH), data_only=True)
    ws = wb[MAIN_SHEET]

    if not dry_run:
        conn = get_db_connection()
    else:
        conn = None
        print("\n=== DRY RUN MODE - No changes will be made ===\n")

    row_num = 0
    for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        row_num += 1

        # Skip empty rows
        if not row[COLUMNS['status']]:
            continue

        # Extract welder data
        status = str(row[COLUMNS['status']]).strip().lower()
        stamp = str(row[COLUMNS['stamp']]) if row[COLUMNS['stamp']] else None
        employee = str(row[COLUMNS['employee']]) if row[COLUMNS['employee']] else None

        # Skip if no identifier
        if not stamp and not employee:
            stats['welders_skipped'] += 1
            continue

        # Clean stamp (remove leading zeros, etc.)
        if stamp and stamp not in ('None', '0', ''):
            stamp = stamp.strip()
        else:
            stamp = None

        # Filter for single welder if specified
        if single_welder:
            if stamp != single_welder and employee != single_welder:
                continue

        stats['welders_processed'] += 1

        # Build welder data dict
        welder_data = {
            'status': 'active' if status == 'active' else 'inactive',
            'welder_stamp': stamp,
            'employee_number': employee if employee not in ('None', '0', 'Rig Welder', 'ULG', 'NCW') else stamp,
            'last_name': str(row[COLUMNS['last_name']]) if row[COLUMNS['last_name']] else '',
            'first_name': str(row[COLUMNS['first_name']]) if row[COLUMNS['first_name']] else '',
            'preferred_name': str(row[COLUMNS['preferred_name']]) if row[COLUMNS['preferred_name']] else None,
            'display_name': str(row[COLUMNS['display_name']]) if row[COLUMNS['display_name']] else None,
            'department': str(row[COLUMNS['department']]) if row[COLUMNS['department']] else None,
            'supervisor': str(row[COLUMNS['supervisor']]) if row[COLUMNS['supervisor']] else None,
            'business_unit': str(row[COLUMNS['business_unit']]) if row[COLUMNS['business_unit']] else None,
            'running_total_welds': int(row[COLUMNS['running_total']] or 0) if COLUMNS['running_total'] < len(row) else 0,
            'total_welds_tested': int(row[COLUMNS['total_tested']] or 0) if COLUMNS['total_tested'] < len(row) else 0,
            'welds_passed': int(row[COLUMNS['pass']] or 0) if COLUMNS['pass'] < len(row) else 0,
            'welds_failed': int(row[COLUMNS['fail']] or 0) if COLUMNS['fail'] < len(row) else 0,
        }

        # Clean None strings
        for key in welder_data:
            if welder_data[key] in ('None', 'none', ''):
                welder_data[key] = None

        row_hash = compute_row_hash(row)

        # Progress reporting
        if stats['welders_processed'] % 50 == 0:
            print(f"  Processed {stats['welders_processed']} welders...")

        if dry_run:
            print(f"[DRY] {welder_data['welder_stamp'] or welder_data['employee_number']}: "
                  f"{welder_data['display_name']} ({welder_data['status']})")
        else:
            # Check if new or update
            existing = None
            if welder_data['welder_stamp']:
                existing = find_welder_by_stamp(conn, welder_data['welder_stamp'])

            welder_id = upsert_welder(conn, welder_data, row_hash)

            if existing:
                if existing.get('excel_row_hash') != row_hash:
                    stats['welders_updated'] += 1
            else:
                stats['welders_created'] += 1

        # Process WPQ codes (columns K-O, indices 10-14)
        wpq_codes = []
        for i in range(COLUMNS['wpq1'], COLUMNS['wpq5'] + 1):
            if i < len(row) and row[i]:
                code = str(row[i]).strip()
                if code and code not in ('', '0', 'None') and not code.isdigit():
                    wpq_codes.append(code)

        for wpq_code in wpq_codes:
            parsed = parse_wpq_code(wpq_code)

            if parsed.get('parse_notes'):
                stats['wpq_parse_errors'] += 1
                stats['errors'].append(f"{welder_data.get('welder_stamp')}: {wpq_code} - {parsed['parse_notes']}")

            if dry_run:
                print(f"    WPQ: {wpq_code} -> {parsed['process_type']} P{parsed.get('p_number')} "
                      f"F{parsed.get('f_number')} {parsed.get('positions', [])}")
            else:
                wpq_id = upsert_wpq(conn, welder_id, wpq_code, parsed)
                if wpq_id:
                    stats['wpqs_created'] += 1

    if conn:
        conn.close()

    return stats


def print_stats(stats: Dict):
    """Print import statistics."""
    print("\n" + "=" * 50)
    print("IMPORT SUMMARY")
    print("=" * 50)
    print(f"Welders processed:  {stats['welders_processed']}")
    print(f"  - Created:        {stats['welders_created']}")
    print(f"  - Updated:        {stats['welders_updated']}")
    print(f"  - Skipped:        {stats['welders_skipped']}")
    print(f"WPQs created:       {stats['wpqs_created']}")
    print(f"Parse warnings:     {stats['wpq_parse_errors']}")

    if stats['errors']:
        print(f"\nParse Issues ({len(stats['errors'])}):")
        for err in stats['errors'][:10]:
            print(f"  - {err}")
        if len(stats['errors']) > 10:
            print(f"  ... and {len(stats['errors']) - 10} more")


# =============================================================================
# MAIN
# =============================================================================

def main():
    parser = argparse.ArgumentParser(description='Import welders from Excel to database')
    parser.add_argument('--dry-run', action='store_true', help='Preview changes without modifying database')
    parser.add_argument('--validate', action='store_true', help='Validate Excel structure only')
    parser.add_argument('--welder', type=str, help='Import single welder by stamp or employee number')

    args = parser.parse_args()

    if args.validate:
        validate_excel()
        return

    if not EXCEL_PATH.exists():
        print(f"ERROR: Excel file not found: {EXCEL_PATH}")
        sys.exit(1)

    stats = import_from_excel(dry_run=args.dry_run, single_welder=args.welder)
    print_stats(stats)


if __name__ == '__main__':
    main()
